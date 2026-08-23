import json
from datetime import datetime, time, timedelta

from django.db import transaction
from django.db.models import F, Max, Prefetch
from django.http import JsonResponse, HttpResponseBadRequest
from django.shortcuts import render, get_object_or_404
from django.utils import timezone
from django.utils.dateparse import parse_datetime, parse_date
from django.views.decorators.csrf import ensure_csrf_cookie
from django.views.decorators.http import require_POST

from accounts.permissions import require_project_role
from django.db.models import Sum

from django.core.exceptions import PermissionDenied

from core.models import (
    ProjectTemplate,
    Baseline, CalendarHoliday, ConstraintType, CustomField, CustomFieldType, DependencyType,
    Methodology, Portfolio, PortfolioProject, Project, ProjectCalendar, ProjectMember, ProjectRole,
    ProjectStatus, RaidItem, RaidPriority, RaidStatus, RaidType, ResourceProfile, ResourceTimeOff,
    SchedulingMode, Sprint, StatusCategory, Task, TaskAssignment, TaskAttachment, TaskComment,
    TaskCustomFieldValue, TaskDependency, TaskStatus, TaskType, TimeEntry, User, WorkflowStatus, Workspace,
    WorkspaceMember, WorkspaceRole,
)
from pmo.notifications import notify_assignment
from pmo.scheduler import (
    DEFAULT_WORKING_DAYS, add_working_days, count_working_days, get_calendar, is_working_day, shift_working_days,
    local_date, next_working_day, required_start_for, reschedule_task_and_successors, to_end_dt,
    to_start_dt,
)

EDIT_ROLES = ["GLOBAL_ADMIN", "OWNER", "EDITOR"]


def _is_admin(request):
    return getattr(request, "pmo_workspace_role", None) == WorkspaceRole.GLOBAL_ADMIN


def _visible_projects(request):
    """Admins see every project in the workspace; everyone else only projects
    they're a member of (change request #13)."""
    qs = Project.objects.filter(workspace=request.pmo_workspace)
    if _is_admin(request):
        return qs
    return qs.filter(memberships__user=request.pmo_user).distinct()


def _require_project_access(request, project):
    """Read access: workspace admin or project member."""
    if _is_admin(request):
        return
    if not ProjectMember.objects.filter(project=project, user=request.pmo_user).exists():
        raise PermissionDenied("You don't have access to this project.")


def _require_admin(request):
    if not _is_admin(request):
        raise PermissionDenied("Only workspace admins can do that.")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _parse_dt(value, end_of_day=False):
    """Accepts ISO datetimes or plain YYYY-MM-DD dates; returns an aware datetime.
    Date-only values are anchored to working hours in the local time zone, so
    they never shift a day when stored as UTC."""
    if not value:
        return None
    value = str(value).strip()
    if len(value) <= 10:  # plain date, e.g. "2026-08-03"
        d = parse_date(value)
        if d is None:
            return None
        dt = datetime.combine(d, time(17, 0) if end_of_day else time(9, 0))
    else:
        dt = parse_datetime(value)
        if dt is None:
            return None
    if timezone.is_naive(dt):
        dt = timezone.make_aware(dt)
    return dt


def _iso(d):
    return d.isoformat() if d else None


def _task_json(t, predecessors, extras=None):
    extras = extras or {}
    return {
        "id": t.id,
        "title": t.title,
        "description": t.description or "",
        "parent_id": t.parent_task_id,
        "status": t.status,
        "progress": t.progress,
        "start": _iso(local_date(t.start_date)),
        "end": _iso(local_date(t.end_date)),
        "sort_order": t.sort_order,
        "is_milestone": t.is_milestone,
        "is_active": t.is_active,
        "task_type": t.task_type,
        "effort_driven": t.effort_driven,
        "scheduling_mode": t.scheduling_mode,
        "fixed_cost": t.fixed_cost,
        "sprint_id": t.sprint_id,
        "assignee": {"id": t.assignee.id, "name": t.assignee.name} if t.assignee else None,
        "assignees": (extras.get("assignments", {}) or {}).get(t.id)
        or ([{"id": t.assignee.id, "name": t.assignee.name, "units": 1.0}] if t.assignee else []),
        "constraint_type": t.constraint_type,
        "constraint_date": _iso(local_date(t.constraint_date)),
        "deadline": _iso(local_date(t.deadline)),
        "estimated_hours": t.estimated_hours,
        "story_points": t.story_points,
        "baseline_start": _iso(local_date(t.baseline_start)),
        "baseline_end": _iso(local_date(t.baseline_end)),
        "logged_minutes": extras.get("logged", {}).get(t.id, 0),
        "custom_values": extras.get("custom", {}).get(t.id, {}),
        "format": t.format_json or {},
        "comments": extras.get("comments", {}).get(t.id, []),
        "attachments": extras.get("attachments", {}).get(t.id, []),
        "predecessors": predecessors.get(t.id, []),
    }


def _predecessor_map(project):
    short = {
        DependencyType.FINISH_TO_START: "FS",
        DependencyType.START_TO_START: "SS",
        DependencyType.FINISH_TO_FINISH: "FF",
        DependencyType.START_TO_FINISH: "SF",
    }
    result = {}
    for dep in TaskDependency.objects.filter(to_task__project=project):
        result.setdefault(dep.to_task_id, []).append(
            {"id": dep.from_task_id, "type": short.get(dep.type, "FS"), "lag": dep.lag_days or 0}
        )
    return result


DEFAULT_STATUSES = [
    ("TODO", "Not Started", StatusCategory.NOT_STARTED, "#94a3b8"),
    ("IN_PROGRESS", "In Progress", StatusCategory.IN_PROGRESS, "#2563eb"),
    ("BLOCKED", "On Hold", StatusCategory.IN_PROGRESS, "#f59e0b"),
    ("DONE", "Complete", StatusCategory.DONE, "#16a34a"),
]


def get_project_statuses(project):
    """This project's statuses, lazily seeding the MS Project-style defaults the
    first time (default keys match legacy Task.status values, so existing tasks
    keep a valid status)."""
    qs = list(project.statuses.all())
    if qs:
        return qs
    return [
        WorkflowStatus.objects.create(project=project, key=key, name=name, category=cat, color=color, sort_order=i)
        for i, (key, name, cat, color) in enumerate(DEFAULT_STATUSES)
    ]


def _status_json(project):
    return [
        {"key": s.key, "name": s.name, "category": s.category, "color": s.color}
        for s in get_project_statuses(project)
    ]


def _status_category_map(project):
    return {s.key: s.category for s in get_project_statuses(project)}


def _resources_json(workspace):
    """Per-person capacity (units) + time off for every workspace member."""
    profiles = {
        p.user_id: p for p in
        ResourceProfile.objects.filter(workspace=workspace).prefetch_related("time_off")
    }
    out = []
    for m in WorkspaceMember.objects.filter(workspace=workspace):
        p = profiles.get(m.user_id)
        out.append({
            "id": m.user_id,
            "units": p.units if p else 1.0,
            "rate": p.rate if p else 0,
            "working_days": (p.working_days_json if (p and p.working_days_json) else None),
            "time_off": [
                {"start": t.start_date.isoformat(), "end": t.end_date.isoformat(), "note": t.note or ""}
                for t in (p.time_off.all() if p else [])
            ],
        })
    return out


def _baselines_json(project):
    return [
        {
            "id": b.id,
            "name": b.name,
            "created_at": timezone.localtime(b.created_at).strftime("%d %b %Y %H:%M"),
            "snapshot": b.snapshot or {},
        }
        for b in project.baselines.all()
    ]


def _slug_status_key(name, taken):
    import re as _re
    base = _re.sub(r"[^a-z0-9]+", "_", (name or "status").strip().lower()).strip("_")[:30] or "status"
    key, n = base, 2
    while key in taken:
        key = f"{base}_{n}"[:40]
        n += 1
    return key


def _calendar_json(project):
    cal = ProjectCalendar.objects.filter(project=project).prefetch_related("holidays").first()
    if cal is None:
        return {"working_days": DEFAULT_WORKING_DAYS, "holidays": [], "time_zone": "Asia/Singapore"}
    return {
        "working_days": cal.working_days_json or DEFAULT_WORKING_DAYS,
        "holidays": [
            {"id": h.id, "name": h.name, "date": timezone.localtime(h.date).date().isoformat()}
            for h in cal.holidays.all().order_by("date")
        ],
        "time_zone": cal.time_zone,
    }


def _would_create_cycle(project_id, from_id, to_id):
    if from_id == to_id:
        return True
    successors = {}
    for dep in TaskDependency.objects.filter(from_task__project_id=project_id):
        successors.setdefault(dep.from_task_id, set()).add(dep.to_task_id)
    seen, stack = set(), [to_id]
    while stack:
        node = stack.pop()
        if node == from_id:
            return True
        if node in seen:
            continue
        seen.add(node)
        stack.extend(successors.get(node, ()))
    return False


def _json_body(request):
    try:
        return json.loads(request.body)
    except (ValueError, json.JSONDecodeError):
        return None


def _portfolio_rollup(portfolio):
    projects = [pp.project for pp in portfolio.portfolio_projects.select_related("project__owner")]
    open_raid = RaidItem.objects.filter(project__in=projects).exclude(status="CLOSED").count() if projects else 0
    starts = [p.start_date for p in projects if p.start_date]
    ends = [p.end_date for p in projects if p.end_date]
    return {
        "portfolio": portfolio,
        "projects": projects,
        "project_count": len(projects),
        "open_raid": open_raid,
        "start": min(starts) if starts else None,
        "end": max(ends) if ends else None,
    }


# ---------------------------------------------------------------------------
# Pages
# ---------------------------------------------------------------------------

def dashboard(request):
    """Workspace dashboard: KPI tiles plus portfolio and project health tables
    (change request #1), respecting per-user visibility (#13)."""
    workspace = request.pmo_workspace
    projects = list(
        _visible_projects(request)
        .select_related("owner")
        .prefetch_related(Prefetch("tasks", queryset=Task.objects.select_related("assignee")), "raid_items")
    )
    visible_ids = [p.id for p in projects]
    today = timezone.now()
    # Any status in the "Complete" category counts as done (custom statuses included).
    done_keys = set(
        WorkflowStatus.objects.filter(project__in=projects, category=StatusCategory.DONE)
        .values_list("key", flat=True)
    ) | {TaskStatus.DONE}
    status_names = {k: n for k, n, _c, _col in DEFAULT_STATUSES}
    status_names.update({s.key: s.name for s in WorkflowStatus.objects.filter(project__in=projects)})

    project_rows = []
    for p in projects:
        tasks = list(p.tasks.all())
        leaf = [t for t in tasks if not any(x.parent_task_id == t.id for x in tasks)]
        done = sum(1 for t in leaf if t.status in done_keys)
        overdue = sum(1 for t in leaf if t.end_date and t.end_date < today and t.status not in done_keys)
        progress = round(sum(t.progress for t in leaf) / len(leaf)) if leaf else 0
        project_rows.append({
            "project": p,
            "progress": progress,
            "task_count": len(leaf),
            "done": done,
            "overdue": overdue,
            "open_raid": sum(1 for r in p.raid_items.all() if r.status != RaidStatus.CLOSED),
            "health": "red" if overdue > 2 else ("amber" if overdue else "green"),
        })

    portfolio_rows = []
    for pf in Portfolio.objects.filter(workspace=workspace).prefetch_related("portfolio_projects__project"):
        pf_projects = [pp.project for pp in pf.portfolio_projects.all() if pp.project_id in visible_ids]
        if not pf_projects and not _is_admin(request):
            continue
        prs = [r for r in project_rows if r["project"].id in {p.id for p in pf_projects}]
        portfolio_rows.append({
            "portfolio": pf,
            "project_count": len(pf_projects),
            "progress": round(sum(r["progress"] for r in prs) / len(prs)) if prs else 0,
            "open_raid": sum(r["open_raid"] for r in prs),
            "overdue": sum(r["overdue"] for r in prs),
        })

    # Overdue and due-soon leaf tasks across visible projects, with assignees.
    soon = today + timedelta(days=7)
    attention = []
    for p in projects:
        tasks = list(p.tasks.all())  # already select_related("assignee") via the prefetch above
        parent_ids = {t.parent_task_id for t in tasks if t.parent_task_id}
        for t in tasks:
            if t.id in parent_ids or not t.end_date or t.status in done_keys:
                continue
            if t.end_date < soon:
                attention.append({
                    "task": t, "project": p,
                    "assignee": t.assignee.name if t.assignee else "Unassigned",
                    "overdue": t.end_date < today,
                    "status_label": status_names.get(t.status, t.status),
                })
    attention.sort(key=lambda r: r["task"].end_date)
    attention = attention[:15]

    kpis = {
        "users": WorkspaceMember.objects.filter(workspace=workspace).count(),
        "projects": len(projects),
        "portfolios": len(portfolio_rows),
        "open_raid": sum(r["open_raid"] for r in project_rows),
    }
    return render(request, "pmo/dashboard.html", {
        "kpis": kpis,
        "portfolio_rows": portfolio_rows,
        "project_rows": project_rows,
        "attention": attention,
        "is_admin": _is_admin(request),
    })


@ensure_csrf_cookie
def portfolio_list(request):
    workspace = request.pmo_workspace
    visible_ids = set(_visible_projects(request).values_list("id", flat=True))
    portfolios = list(Portfolio.objects.filter(workspace=workspace).select_related("owner").order_by("name"))

    def build(p):
        row = _portfolio_rollup(p)
        row["projects"] = [pr for pr in row["projects"] if pr.id in visible_ids]
        kids = [k for k in (build(c) for c in portfolios if c.parent_id == p.id) if k]
        row["children"] = kids
        row["project_count"] = len(row["projects"]) + sum(k["project_count"] for k in kids)
        row["open_raid"] = row["open_raid"] + sum(k["open_raid"] for k in kids)
        if not row["projects"] and not kids and not _is_admin(request):
            return None
        return row

    rows = [r for r in (build(p) for p in portfolios if p.parent_id is None) if r]
    assigned_ids = PortfolioProject.objects.filter(portfolio__workspace=workspace).values_list("project_id", flat=True)
    unassigned = Project.objects.filter(workspace=workspace, id__in=visible_ids).exclude(id__in=assigned_ids).order_by("name")
    return render(request, "pmo/portfolio_list.html", {
        "rows": rows,
        "unassigned": unassigned,
        "all_portfolios": portfolios,
        "status_choices": ProjectStatus.choices,
    })


@ensure_csrf_cookie
def project_list(request):
    workspace = request.pmo_workspace
    portfolios = Portfolio.objects.filter(workspace=workspace).order_by("name")
    link_map = {}
    for pp in PortfolioProject.objects.filter(portfolio__workspace=workspace).select_related("portfolio"):
        link_map.setdefault(pp.project_id, pp.portfolio)

    projects = _visible_projects(request).select_related("owner").prefetch_related(
        "tasks", "raid_items"
    ).order_by("-created_at")
    rows = [
        {
            "project": p,
            "portfolio": link_map.get(p.id),
            "task_count": p.tasks.count(),
            "open_raid_count": p.raid_items.exclude(status="CLOSED").count(),
        }
        for p in projects
    ]
    return render(request, "pmo/project_list.html", {
        "rows": rows,
        "portfolios": portfolios,
        "templates": ProjectTemplate.objects.filter(workspace=workspace).order_by("name"),
        "status_choices": ProjectStatus.choices,
    })


@ensure_csrf_cookie
def project_detail(request, project_id):
    """React/TypeScript planner (local-first editing, undo/redo)."""
    project = get_object_or_404(Project, id=project_id, workspace=request.pmo_workspace)
    _require_project_access(request, project)
    link = PortfolioProject.objects.filter(project=project).select_related("portfolio").first()
    import os as _os
    try:
        _pv = int(_os.path.getmtime(_os.path.join(_os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))), "static", "pmo", "planner", "planner.js")))
    except OSError:
        _pv = 1
    return render(request, "pmo/project_plan.html", {
        "project": project,
        "planner_v": _pv,
        "portfolio": link.portfolio if link else None,
        "task_statuses_json": json.dumps([[s["key"], s["name"]] for s in _status_json(project)]),
    })


@ensure_csrf_cookie
def project_classic(request, project_id):
    """The previous server-rendered editor, kept as a fallback."""
    project = get_object_or_404(
        Project.objects.select_related("owner").prefetch_related("tasks", "raid_items", "memberships", "time_entries"),
        id=project_id, workspace=request.pmo_workspace,
    )
    _require_project_access(request, project)
    link = PortfolioProject.objects.filter(project=project).select_related("portfolio").first()
    tasks = list(project.tasks.all())
    open_raid = project.raid_items.exclude(status="CLOSED").count()
    avg_progress = round(sum(t.progress for t in tasks) / len(tasks)) if tasks else 0
    return render(request, "pmo/project_detail.html", {
        "project": project,
        "portfolio": link.portfolio if link else None,
        "open_raid": open_raid,
        "avg_progress": avg_progress,
        "member_count": project.memberships.count(),
        "status_choices": ProjectStatus.choices,
        "task_statuses": TaskStatus.choices,
    })


@ensure_csrf_cookie
def project_board(request, project_id):
    """Agile board: one column per task status, drag cards to move them."""
    project = get_object_or_404(Project, id=project_id, workspace=request.pmo_workspace)
    _require_project_access(request, project)
    link = PortfolioProject.objects.filter(project=project).select_related("portfolio").first()
    return render(request, "pmo/project_board.html", {
        "project": project,
        "portfolio": link.portfolio if link else None,
        "task_statuses": TaskStatus.choices,
    })


@ensure_csrf_cookie
def raid_list(request, project_id):
    project = get_object_or_404(Project, id=project_id, workspace=request.pmo_workspace)
    _require_project_access(request, project)
    items = project.raid_items.select_related("owner").order_by("-priority", "status")
    members = WorkspaceMember.objects.filter(workspace=request.pmo_workspace).select_related("user")
    return render(request, "pmo/raid_list.html", {
        "project": project, "items": items,
        "members": [m.user for m in members],
        "raid_types": RaidType.choices, "raid_statuses": RaidStatus.choices,
        "raid_priorities": RaidPriority.choices,
    })


def portfolio_raid(request, portfolio_id):
    """RAID log rolled up across every project in a portfolio (change request #14)."""
    portfolio = get_object_or_404(Portfolio, id=portfolio_id, workspace=request.pmo_workspace)
    visible_ids = set(_visible_projects(request).values_list("id", flat=True))
    projects = [pp.project for pp in portfolio.portfolio_projects.select_related("project")
                if pp.project_id in visible_ids]
    items = RaidItem.objects.filter(project__in=projects).select_related("project", "owner") \
                            .order_by("project__name", "-priority", "status")
    return render(request, "pmo/portfolio_raid.html", {"portfolio": portfolio, "items": items})


def project_dashboard(request, project_id):
    """Steerco-level dashboard for one project (distinct from the workspace dashboard)."""
    project = get_object_or_404(Project, id=project_id, workspace=request.pmo_workspace)
    _require_project_access(request, project)
    tasks = list(project.tasks.all())
    parent_ids = {t.parent_task_id for t in tasks if t.parent_task_id}
    leaf = [t for t in tasks if t.id not in parent_ids]
    now = timezone.now()

    def pct(part, whole):
        return round(part / whole * 100) if whole else 0

    done = [t for t in leaf if t.status == TaskStatus.DONE]
    overdue = [t for t in leaf if t.end_date and t.end_date < now and t.status != TaskStatus.DONE]
    milestones = [t for t in leaf if t.is_milestone]
    ms_done = [t for t in milestones if t.status == TaskStatus.DONE]
    progress = round(sum(t.progress for t in leaf) / len(leaf)) if leaf else 0

    starts = [t.start_date for t in leaf if t.start_date]
    ends = [t.end_date for t in leaf if t.end_date]
    p_start = min(starts) if starts else None
    p_end = max(ends) if ends else None
    elapsed_pct = 0
    if p_start and p_end and p_end > p_start:
        elapsed_pct = max(0, min(100, round((now - p_start).total_seconds() / (p_end - p_start).total_seconds() * 100)))

    status_slices = []
    palette = {"TODO": "#94a3b8", "IN_PROGRESS": "#0033a0", "BLOCKED": "#f59e0b", "DONE": "#16a34a"}
    for value, label in TaskStatus.choices:
        n = sum(1 for t in leaf if t.status == value)
        if n:
            status_slices.append({"label": label, "n": n, "pct": pct(n, len(leaf)), "color": palette.get(value, "#64748b")})

    by_id = {t.id: t for t in tasks}

    def under(t, ancestor_id):
        node = t
        while node is not None:
            if node.id == ancestor_id:
                return True
            node = by_id.get(node.parent_task_id)
        return False

    phases = []
    for ph in [t for t in tasks if t.parent_task_id is None]:
        sub = [t for t in leaf if under(t, ph.id)]
        if sub:
            phases.append({"name": ph.title,
                           "progress": round(sum(t.progress for t in sub) / len(sub)),
                           "done": sum(1 for t in sub if t.status == TaskStatus.DONE), "total": len(sub)})

    upcoming_ms = sorted([t for t in milestones if t.status != TaskStatus.DONE and t.end_date],
                         key=lambda t: t.end_date)[:8]
    raid_top = project.raid_items.exclude(status=RaidStatus.CLOSED).order_by("-priority")[:8]
    workload = {}
    for t in leaf:
        key = t.assignee.name if t.assignee else "Unassigned"
        w = workload.setdefault(key, {"total": 0, "done": 0, "overdue": 0})
        w["total"] += 1
        if t.status == TaskStatus.DONE:
            w["done"] += 1
        if t.end_date and t.end_date < now and t.status != TaskStatus.DONE:
            w["overdue"] += 1
    workload_rows = sorted(({"name": k, **v} for k, v in workload.items()), key=lambda r: -r["total"])[:8]

    spi_hint = "on track" if progress >= elapsed_pct - 5 else ("at risk" if progress >= elapsed_pct - 15 else "behind")
    return render(request, "pmo/project_dashboard.html", {
        "project": project,
        "progress": progress, "elapsed_pct": elapsed_pct, "spi_hint": spi_hint,
        "leaf_count": len(leaf), "done_count": len(done), "overdue": overdue[:8], "overdue_count": len(overdue),
        "ms_total": len(milestones), "ms_done": len(ms_done),
        "open_raid": project.raid_items.exclude(status=RaidStatus.CLOSED).count(),
        "p_start": p_start, "p_end": p_end,
        "status_slices": status_slices, "phases": phases,
        "upcoming_ms": upcoming_ms, "raid_top": raid_top, "workload_rows": workload_rows,
    })


def raid_overview(request):
    """All open RAID items across every project the user can see (dashboard link)."""
    projects = list(_visible_projects(request))
    items = RaidItem.objects.filter(project__in=projects).exclude(status=RaidStatus.CLOSED) \
                            .select_related("project", "owner").order_by("-priority", "project__name")
    return render(request, "pmo/raid_overview.html", {"items": items})


def project_timelog(request, project_id):
    """Per-project time log with a weekly bandwidth matrix per team member (change #6)."""
    project = get_object_or_404(Project, id=project_id, workspace=request.pmo_workspace)
    _require_project_access(request, project)

    entries = list(project.time_entries.select_related("user", "task").order_by("-entry_date")[:200])

    # Weekly bandwidth: hours per user for the last 4 ISO weeks + total.
    now = timezone.localtime(timezone.now())
    week_starts = []
    monday = (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
    for i in range(3, -1, -1):
        week_starts.append(monday - timedelta(weeks=i))

    matrix = {}  # user -> {"weeks": [h,h,h,h], "total": h}
    for e in project.time_entries.select_related("user"):
        row = matrix.setdefault(e.user, {"weeks": [0.0] * 4, "total": 0.0})
        hours = e.minutes / 60
        row["total"] += hours
        ed = timezone.localtime(e.entry_date)
        for i, ws in enumerate(week_starts):
            if ws <= ed < ws + timedelta(weeks=1):
                row["weeks"][i] += hours
                break
    bandwidth = [
        {"user": u, "weeks": [round(h, 1) for h in d["weeks"]], "total": round(d["total"], 1),
         "recent": round(sum(d["weeks"]), 1)}
        for u, d in sorted(matrix.items(), key=lambda kv: -kv[1]["total"])
    ]
    return render(request, "pmo/project_timelog.html", {
        "project": project,
        "entries": entries,
        "bandwidth": bandwidth,
        "week_labels": [ws.strftime("%d %b") for ws in week_starts],
    })


@require_POST
def raid_create(request, project_id):
    require_project_role(request, project_id, allowed_roles=EDIT_ROLES)
    project = get_object_or_404(Project, id=project_id, workspace=request.pmo_workspace)
    payload = _json_body(request)
    if payload is None or not (payload.get("title") or "").strip():
        return HttpResponseBadRequest("A title is required.")
    item = RaidItem.objects.create(
        project=project,
        type=payload.get("type") if payload.get("type") in RaidType.values else RaidType.RISK,
        title=payload["title"].strip(),
        description=(payload.get("description") or "").strip() or None,
        priority=payload.get("priority") if payload.get("priority") in RaidPriority.values else RaidPriority.MEDIUM,
        status=payload.get("status") if payload.get("status") in RaidStatus.values else RaidStatus.OPEN,
        owner_id=payload.get("owner_id") or None,
        due_date=_parse_dt(payload.get("due_date"), end_of_day=True),
        mitigation_plan=(payload.get("mitigation_plan") or "").strip() or None,
    )
    return JsonResponse({"ok": True, "id": item.id})


@require_POST
def raid_update(request, project_id, item_id):
    require_project_role(request, project_id, allowed_roles=EDIT_ROLES)
    item = get_object_or_404(RaidItem, id=item_id, project_id=project_id,
                             project__workspace=request.pmo_workspace)
    payload = _json_body(request)
    if payload is None:
        return HttpResponseBadRequest("Expected a JSON body.")
    if "title" in payload and (payload["title"] or "").strip():
        item.title = payload["title"].strip()
    if "type" in payload and payload["type"] in RaidType.values:
        item.type = payload["type"]
    if "priority" in payload and payload["priority"] in RaidPriority.values:
        item.priority = payload["priority"]
    if "status" in payload and payload["status"] in RaidStatus.values:
        item.status = payload["status"]
    if "owner_id" in payload:
        item.owner_id = payload["owner_id"] or None
    if "due_date" in payload:
        item.due_date = _parse_dt(payload["due_date"], end_of_day=True)
    if "description" in payload:
        item.description = (payload["description"] or "").strip() or None
    if "mitigation_plan" in payload:
        item.mitigation_plan = (payload["mitigation_plan"] or "").strip() or None
    if "resolution_note" in payload:
        item.resolution_note = (payload["resolution_note"] or "").strip() or None
    item.save()
    return JsonResponse({"ok": True})


@require_POST
def raid_delete(request, project_id, item_id):
    require_project_role(request, project_id, allowed_roles=EDIT_ROLES)
    item = get_object_or_404(RaidItem, id=item_id, project_id=project_id,
                             project__workspace=request.pmo_workspace)
    item.delete()
    return JsonResponse({"ok": True})


@ensure_csrf_cookie
def team(request):
    members = WorkspaceMember.objects.filter(workspace=request.pmo_workspace).select_related("user")
    rows = []
    for m in members:
        project_count = ProjectMember.objects.filter(user=m.user).count()
        total_minutes = sum(e.minutes for e in m.user.time_entries.all())
        rows.append({
            "id": m.user.id,
            "name": m.user.name,
            "email": m.user.email,
            "workspace_role": m.role,
            "project_count": project_count,
            "total_hours": round(total_minutes / 60, 1),
            "is_me": m.user.id == request.pmo_user.id,
        })
    return render(request, "pmo/team.html", {
        "rows": rows,
        "is_admin": _is_admin(request),
        "roles": WorkspaceRole.choices,
        "projects": _visible_projects(request).order_by("name"),
        "project_roles": ProjectRole.choices,
    })


@require_POST
def user_add(request):
    """Admin adds a user to the workspace (Entra ID signs them in on first visit)."""
    _require_admin(request)
    payload = _json_body(request)
    if payload is None or not (payload.get("email") or "").strip():
        return HttpResponseBadRequest("An email address is required.")
    email = payload["email"].strip().lower()
    user, _ = User.objects.get_or_create(
        email=email, defaults={"name": (payload.get("name") or "").strip() or email}
    )
    role = payload.get("role") if payload.get("role") in WorkspaceRole.values else WorkspaceRole.MEMBER
    WorkspaceMember.objects.update_or_create(
        workspace=request.pmo_workspace, user=user, defaults={"role": role}
    )
    # Admins see the whole workspace — no per-project grant needed (change #5).
    if role != WorkspaceRole.GLOBAL_ADMIN and payload.get("project_id"):
        project = get_object_or_404(Project, id=payload["project_id"], workspace=request.pmo_workspace)
        p_role = payload.get("project_role") if payload.get("project_role") in ProjectRole.values else ProjectRole.EDITOR
        ProjectMember.objects.update_or_create(project=project, user=user, defaults={"role": p_role})
    return JsonResponse({"ok": True, "id": user.id})


@require_POST
def user_role(request, user_id):
    _require_admin(request)
    payload = _json_body(request)
    role = (payload or {}).get("role")
    if role not in WorkspaceRole.values:
        return HttpResponseBadRequest("role must be GLOBAL_ADMIN or MEMBER.")
    member = get_object_or_404(WorkspaceMember, workspace=request.pmo_workspace, user_id=user_id)
    if member.user_id == request.pmo_user.id and role != WorkspaceRole.GLOBAL_ADMIN:
        return HttpResponseBadRequest("You can't demote yourself.")
    member.role = role
    member.save()
    return JsonResponse({"ok": True})


@require_POST
def user_remove(request, user_id):
    _require_admin(request)
    if user_id == request.pmo_user.id:
        return HttpResponseBadRequest("You can't remove yourself.")
    member = get_object_or_404(WorkspaceMember, workspace=request.pmo_workspace, user_id=user_id)
    ProjectMember.objects.filter(user_id=user_id, project__workspace=request.pmo_workspace).delete()
    member.delete()
    return JsonResponse({"ok": True})


@require_POST
def user_assign_project(request):
    """Admin grants a user a role on a project (drives visibility for non-admins)."""
    _require_admin(request)
    payload = _json_body(request) or {}
    project = get_object_or_404(Project, id=payload.get("project_id"), workspace=request.pmo_workspace)
    user = get_object_or_404(User, id=payload.get("user_id"))
    role = payload.get("role") if payload.get("role") in ProjectRole.values else ProjectRole.EDITOR
    if payload.get("remove"):
        ProjectMember.objects.filter(project=project, user=user).delete()
    else:
        ProjectMember.objects.update_or_create(project=project, user=user, defaults={"role": role})
    return JsonResponse({"ok": True})


# ---------------------------------------------------------------------------
# Portfolio API
# ---------------------------------------------------------------------------

@require_POST
def portfolio_create(request):
    payload = _json_body(request)
    if payload is None or not (payload.get("name") or "").strip():
        return HttpResponseBadRequest("A portfolio name is required.")
    portfolio = Portfolio.objects.create(
        workspace=request.pmo_workspace,
        name=payload["name"].strip(),
        description=(payload.get("description") or "").strip() or None,
        owner=request.pmo_user,
    )
    return JsonResponse({"ok": True, "id": portfolio.id})


@require_POST
def portfolio_delete(request, portfolio_id):
    portfolio = get_object_or_404(Portfolio, id=portfolio_id, workspace=request.pmo_workspace)
    portfolio.delete()  # link rows cascade; projects themselves are kept
    return JsonResponse({"ok": True})


@require_POST
def portfolio_assign(request):
    """Move a project into a portfolio (or out of all portfolios with portfolio_id: null)."""
    payload = _json_body(request)
    if payload is None or "project_id" not in payload:
        return HttpResponseBadRequest("Expected project_id and portfolio_id.")
    project = get_object_or_404(Project, id=payload["project_id"], workspace=request.pmo_workspace)
    with transaction.atomic():
        PortfolioProject.objects.filter(project=project).delete()
        if payload.get("portfolio_id"):
            portfolio = get_object_or_404(Portfolio, id=payload["portfolio_id"], workspace=request.pmo_workspace)
            PortfolioProject.objects.create(portfolio=portfolio, project=project)
    return JsonResponse({"ok": True})


# ---------------------------------------------------------------------------
# Project API
# ---------------------------------------------------------------------------

def projects_json(request):
    """Visible projects as JSON (feeds the planner's File → Open dialog)."""
    return JsonResponse({
        "projects": [
            {"id": p.id, "name": p.name, "status": p.get_status_display()}
            for p in _visible_projects(request).order_by("name")
        ]
    })



# ---------------------------------------------------------------------------
# Project templates (save a plan as a template, create projects from it)
# ---------------------------------------------------------------------------

_DEP_SHORT = {"FINISH_TO_START": "FS", "START_TO_START": "SS", "FINISH_TO_FINISH": "FF", "START_TO_FINISH": "SF"}
_DEP_LONG = {v: k for k, v in _DEP_SHORT.items()}


@ensure_csrf_cookie
def templates_page(request):
    """Template gallery: saved templates + built-in starters (Smartsheet-style)."""
    tpls = ProjectTemplate.objects.filter(workspace=request.pmo_workspace).order_by("-created_at")
    return render(request, "pmo/templates_page.html", {
        "tpls": tpls,
        "counts": {t.id: len(t.data_json.get("tasks", [])) for t in tpls},
    })


BUILTIN_TEMPLATES = {
    "software": {"name": "Software Delivery (starter)", "description": "Discovery → Build → Test → Deploy with milestones", "tasks": [
        {"i": 0, "title": "Discovery & Requirements", "parent_i": None, "start_off": 0, "dur": 1, "milestone": False, "preds": []},
        {"i": 1, "title": "Stakeholder workshops", "parent_i": 0, "start_off": 0, "dur": 5, "milestone": False, "preds": []},
        {"i": 2, "title": "Requirements sign-off", "parent_i": 0, "start_off": 5, "dur": 1, "milestone": True, "preds": [{"i": 1, "type": "FS", "lag": 0}]},
        {"i": 3, "title": "Design & Build", "parent_i": None, "start_off": 6, "dur": 1, "milestone": False, "preds": []},
        {"i": 4, "title": "Solution design", "parent_i": 3, "start_off": 6, "dur": 5, "milestone": False, "preds": [{"i": 2, "type": "FS", "lag": 0}]},
        {"i": 5, "title": "Development", "parent_i": 3, "start_off": 11, "dur": 15, "milestone": False, "preds": [{"i": 4, "type": "FS", "lag": 0}]},
        {"i": 6, "title": "Test & Deploy", "parent_i": None, "start_off": 26, "dur": 1, "milestone": False, "preds": []},
        {"i": 7, "title": "System testing", "parent_i": 6, "start_off": 26, "dur": 7, "milestone": False, "preds": [{"i": 5, "type": "FS", "lag": 0}]},
        {"i": 8, "title": "UAT", "parent_i": 6, "start_off": 33, "dur": 5, "milestone": False, "preds": [{"i": 7, "type": "FS", "lag": 0}]},
        {"i": 9, "title": "Go-live", "parent_i": 6, "start_off": 38, "dur": 1, "milestone": True, "preds": [{"i": 8, "type": "FS", "lag": 0}]},
    ]},
    "rollout": {"name": "Client Rollout (starter)", "description": "Kickoff → Configure → Migrate → Hypercare", "tasks": [
        {"i": 0, "title": "Kickoff", "parent_i": None, "start_off": 0, "dur": 1, "milestone": True, "preds": []},
        {"i": 1, "title": "Configuration", "parent_i": None, "start_off": 1, "dur": 10, "milestone": False, "preds": [{"i": 0, "type": "FS", "lag": 0}]},
        {"i": 2, "title": "Data migration", "parent_i": None, "start_off": 11, "dur": 5, "milestone": False, "preds": [{"i": 1, "type": "FS", "lag": 0}]},
        {"i": 3, "title": "Training", "parent_i": None, "start_off": 11, "dur": 5, "milestone": False, "preds": [{"i": 1, "type": "SS", "lag": 0}]},
        {"i": 4, "title": "Cutover", "parent_i": None, "start_off": 16, "dur": 2, "milestone": False, "preds": [{"i": 2, "type": "FS", "lag": 0}, {"i": 3, "type": "FS", "lag": 0}]},
        {"i": 5, "title": "Hypercare complete", "parent_i": None, "start_off": 23, "dur": 1, "milestone": True, "preds": [{"i": 4, "type": "FS", "lag": 5}]},
    ]},
}


@require_POST
def template_add_builtin(request):
    payload = _json_body(request) or {}
    kind = payload.get("kind")
    if kind not in BUILTIN_TEMPLATES:
        return HttpResponseBadRequest("Unknown starter template.")
    spec = BUILTIN_TEMPLATES[kind]
    tpl = ProjectTemplate.objects.create(
        workspace=request.pmo_workspace, name=spec["name"], description=spec["description"],
        data_json={"tasks": spec["tasks"]}, created_by=request.pmo_user,
    )
    return JsonResponse({"ok": True, "id": tpl.id})


def templates_json(request):
    return JsonResponse({"templates": [
        {"id": t.id, "name": t.name, "description": t.description or "",
         "tasks": len(t.data_json.get("tasks", [])),
         "created": timezone.localtime(t.created_at).strftime("%d %b %Y"),
         "by": t.created_by.name if t.created_by else ""}
        for t in ProjectTemplate.objects.filter(workspace=request.pmo_workspace).order_by("-created_at")
    ]})


@require_POST
def template_save_from(request, project_id):
    """Snapshot a project's task structure (relative working-day offsets) as a template."""
    project = get_object_or_404(Project, id=project_id, workspace=request.pmo_workspace)
    _require_project_access(request, project)
    payload = _json_body(request)
    if payload is None or not (payload.get("name") or "").strip():
        return HttpResponseBadRequest("A template name is required.")

    cal = get_calendar(project)
    tasks = list(project.tasks.all())
    index_of = {t.id: i for i, t in enumerate(tasks)}
    starts = [local_date(t.start_date) for t in tasks if t.start_date]
    base = min(starts) if starts else timezone.localtime(timezone.now()).date()
    deps = {}
    for dep in TaskDependency.objects.filter(to_task__project=project):
        deps.setdefault(dep.to_task_id, []).append({
            "i": index_of.get(dep.from_task_id), "type": _DEP_SHORT.get(dep.type, "FS"), "lag": dep.lag_days or 0,
        })
    rows = []
    for i, t in enumerate(tasks):
        s0 = local_date(t.start_date)
        rows.append({
            "i": i, "title": t.title, "description": t.description or "",
            "parent_i": index_of.get(t.parent_task_id),
            "start_off": max(0, count_working_days(base, s0, cal) - 1) if s0 else 0,
            "dur": count_working_days(local_date(t.start_date), local_date(t.end_date), cal) if t.start_date and t.end_date else 1,
            "milestone": t.is_milestone,
            "preds": [p for p in deps.get(t.id, []) if p["i"] is not None],
        })
    tpl = ProjectTemplate.objects.create(
        workspace=request.pmo_workspace, name=payload["name"].strip(),
        description=(payload.get("description") or "").strip() or None,
        data_json={"tasks": rows}, created_by=request.pmo_user,
    )
    return JsonResponse({"ok": True, "id": tpl.id, "tasks": len(rows)})


@require_POST
def template_delete(request, template_id):
    tpl = get_object_or_404(ProjectTemplate, id=template_id, workspace=request.pmo_workspace)
    tpl.delete()
    return JsonResponse({"ok": True})


def _instantiate_template(project, tpl):
    cal = get_calendar(project)
    base = local_date(project.start_date) if project.start_date else timezone.localtime(timezone.now()).date()
    base = next_working_day(base, cal)
    id_by_i, pending = {}, []
    for row in tpl.data_json.get("tasks", []):
        start = shift_working_days(base, int(row.get("start_off", 0)), cal)
        end = add_working_days(start, max(1, int(row.get("dur", 1))), cal)
        parent_i = row.get("parent_i")
        task = Task.objects.create(
            project=project,
            parent_task_id=id_by_i.get(parent_i) if parent_i is not None else None,
            title=(row.get("title") or "Task")[:500],
            description=row.get("description") or None,
            start_date=to_start_dt(start),
            end_date=to_start_dt(start) if row.get("milestone") else to_end_dt(end),
            sort_order=int(row.get("i", 0)) + 1,
            is_milestone=bool(row.get("milestone")),
        )
        id_by_i[row.get("i")] = task.id
        for pr in row.get("preds", []):
            pending.append((task.id, pr))
    for tid, pr in pending:
        frm = id_by_i.get(pr.get("i"))
        if frm and frm != tid:
            TaskDependency.objects.create(
                from_task_id=frm, to_task_id=tid,
                type=_DEP_LONG.get((pr.get("type") or "FS").upper(), "FINISH_TO_START"),
                lag_days=int(pr.get("lag") or 0),
            )
    return len(id_by_i)


@require_POST
def project_create(request):
    payload = _json_body(request)
    if payload is None or not (payload.get("name") or "").strip():
        return HttpResponseBadRequest("A project name is required.")

    with transaction.atomic():
        project = Project.objects.create(
            workspace=request.pmo_workspace,
            name=payload["name"].strip(),
            description=(payload.get("description") or "").strip() or None,
            status=payload.get("status") or ProjectStatus.PLANNING,
            owner=request.pmo_user,
            start_date=_parse_dt(payload.get("start_date")),
            end_date=_parse_dt(payload.get("end_date"), end_of_day=True),
        )
        ProjectMember.objects.create(project=project, user=request.pmo_user, role=ProjectRole.OWNER)
        ProjectCalendar.objects.create(project=project, working_days_json=DEFAULT_WORKING_DAYS)
        if payload.get("portfolio_id"):
            portfolio = get_object_or_404(Portfolio, id=payload["portfolio_id"], workspace=request.pmo_workspace)
            PortfolioProject.objects.create(portfolio=portfolio, project=project)
    if payload.get("template_id"):
        tpl = get_object_or_404(ProjectTemplate, id=payload["template_id"], workspace=request.pmo_workspace)
        _instantiate_template(project, tpl)
    return JsonResponse({"ok": True, "id": project.id, "url": f"/projects/{project.id}/"})


@require_POST
def project_update(request, project_id):
    require_project_role(request, project_id, allowed_roles=EDIT_ROLES)
    project = get_object_or_404(Project, id=project_id, workspace=request.pmo_workspace)
    payload = _json_body(request)
    if payload is None:
        return HttpResponseBadRequest("Expected a JSON body.")

    if "name" in payload:
        if not (payload["name"] or "").strip():
            return HttpResponseBadRequest("A project name is required.")
        project.name = payload["name"].strip()
    if "description" in payload:
        project.description = (payload["description"] or "").strip() or None
    if "status" in payload and payload["status"] in ProjectStatus.values:
        project.status = payload["status"]
    if "methodology" in payload and payload["methodology"] in Methodology.values:
        project.methodology = payload["methodology"]
    if "start_date" in payload:
        project.start_date = _parse_dt(payload["start_date"])
    if "end_date" in payload:
        project.end_date = _parse_dt(payload["end_date"], end_of_day=True)
    project.save()
    return JsonResponse({"ok": True})


# ---------------------------------------------------------------------------
# Calendar API (MS Project-style project calendar)
# ---------------------------------------------------------------------------

@require_POST
def calendar_update(request, project_id):
    """Replace the project's working-day pattern and holiday list."""
    require_project_role(request, project_id, allowed_roles=EDIT_ROLES)
    project = get_object_or_404(Project, id=project_id, workspace=request.pmo_workspace)
    payload = _json_body(request)
    if payload is None:
        return HttpResponseBadRequest("Expected a JSON body.")

    working_days = payload.get("working_days")
    if not isinstance(working_days, list) or not working_days or \
            not all(isinstance(d, int) and 1 <= d <= 7 for d in working_days):
        return HttpResponseBadRequest("working_days must be a non-empty list of ISO weekday numbers (1=Mon..7=Sun).")

    with transaction.atomic():
        cal, _ = ProjectCalendar.objects.get_or_create(
            project=project, defaults={"working_days_json": DEFAULT_WORKING_DAYS}
        )
        cal.working_days_json = sorted(set(working_days))
        if payload.get("time_zone"):
            cal.time_zone = payload["time_zone"]
        cal.save()

        if "holidays" in payload:
            cal.holidays.all().delete()
            for h in payload["holidays"] or []:
                d = _parse_dt(h.get("date"))
                if d is None or not (h.get("name") or "").strip():
                    return HttpResponseBadRequest("Each holiday needs a name and a date.")
                CalendarHoliday.objects.create(calendar=cal, name=h["name"].strip(), date=d)

    return JsonResponse({"ok": True, "calendar": _calendar_json(project)})


# ---------------------------------------------------------------------------
# Sprint API (versions/iterations for the agile board)
# ---------------------------------------------------------------------------

@require_POST
def sprint_create(request, project_id):
    require_project_role(request, project_id, allowed_roles=EDIT_ROLES)
    project = get_object_or_404(Project, id=project_id, workspace=request.pmo_workspace)
    payload = _json_body(request)
    if payload is None or not (payload.get("name") or "").strip():
        return HttpResponseBadRequest("A sprint name is required.")
    sprint = Sprint.objects.create(
        project=project,
        name=payload["name"].strip(),
        start_date=_parse_dt(payload.get("start_date")),
        end_date=_parse_dt(payload.get("end_date"), end_of_day=True),
        sort_order=(project.sprints.aggregate(m=Max("sort_order"))["m"] or 0) + 1,
    )
    return JsonResponse({"ok": True, "id": sprint.id})


@require_POST
def sprint_update(request, project_id, sprint_id):
    require_project_role(request, project_id, allowed_roles=EDIT_ROLES)
    sprint = get_object_or_404(Sprint, id=sprint_id, project_id=project_id,
                               project__workspace=request.pmo_workspace)
    payload = _json_body(request)
    if payload is None:
        return HttpResponseBadRequest("Expected a JSON body.")
    if "name" in payload and (payload["name"] or "").strip():
        sprint.name = payload["name"].strip()
    if "start_date" in payload:
        sprint.start_date = _parse_dt(payload["start_date"])
    if "end_date" in payload:
        sprint.end_date = _parse_dt(payload["end_date"], end_of_day=True)
    sprint.save()
    return JsonResponse({"ok": True})


@require_POST
def sprint_delete(request, project_id, sprint_id):
    """Deleting a sprint sends its tasks back to the backlog (sprint = null)."""
    require_project_role(request, project_id, allowed_roles=EDIT_ROLES)
    sprint = get_object_or_404(Sprint, id=sprint_id, project_id=project_id,
                               project__workspace=request.pmo_workspace)
    sprint.delete()  # Task.sprint is SET_NULL
    return JsonResponse({"ok": True})


# ---------------------------------------------------------------------------
# Task API
# ---------------------------------------------------------------------------

def project_tasks_data(request, project_id):
    """Everything the grid + Gantt editor needs, in one payload."""
    project = get_object_or_404(Project, id=project_id, workspace=request.pmo_workspace)
    _require_project_access(request, project)
    predecessors = _predecessor_map(project)
    logged = dict(
        TimeEntry.objects.filter(project=project, task__isnull=False)
        .values_list("task_id").annotate(total=Sum("minutes")).values_list("task_id", "total")
    )
    custom = {}
    for v in TaskCustomFieldValue.objects.filter(task__project=project):
        custom.setdefault(v.task_id, {})[v.custom_field_id] = v.value
    comments = {}
    for c in TaskComment.objects.filter(task__project=project).select_related("user"):
        comments.setdefault(c.task_id, []).append({
            "id": c.id, "user": c.user.name, "body": c.body,
            "at": timezone.localtime(c.created_at).strftime("%d %b %Y %H:%M"),
        })
    attachments = {}
    for a in TaskAttachment.objects.filter(task__project=project):
        kind = "file" if (a.mime_type and a.mime_type != "link") else "link"
        attachments.setdefault(a.task_id, []).append(
            {"id": a.id, "name": a.file_name, "url": a.blob_url, "mime": a.mime_type or "", "kind": kind})
    assignments = {}
    for asg in TaskAssignment.objects.filter(task__project=project).select_related("user"):
        assignments.setdefault(asg.task_id, []).append({"id": asg.user_id, "name": asg.user.name, "units": asg.units})
    extras = {"logged": logged, "custom": custom, "comments": comments,
              "attachments": attachments, "assignments": assignments}
    return JsonResponse({
        "project": {
            "id": project.id,
            "name": project.name,
            "methodology": project.methodology,
            "start": local_date(project.start_date).isoformat() if project.start_date else None,
            "end": local_date(project.end_date).isoformat() if project.end_date else None,
        },
        "calendar": _calendar_json(project),
        "sprints": [
            {
                "id": s.id,
                "name": s.name,
                "start": local_date(s.start_date).isoformat() if s.start_date else None,
                "end": local_date(s.end_date).isoformat() if s.end_date else None,
            }
            for s in project.sprints.all()
        ],
        "members": [
            {"id": m.user.id, "name": m.user.name, "email": m.user.email}
            for m in WorkspaceMember.objects.filter(workspace=request.pmo_workspace)
                                            .select_related("user").order_by("user__name")
        ],
        "custom_fields": [
            {"id": f.id, "name": f.name, "type": f.type, "options": f.options_json or []}
            for f in project.custom_fields.all().order_by("id")
        ],
        "statuses": _status_json(project),
        "baselines": _baselines_json(project),
        "resources": _resources_json(request.pmo_workspace),
        "tasks": [_task_json(t, predecessors, extras) for t in project.tasks.select_related("assignee")],
    })


@require_POST
def task_create(request, project_id):
    require_project_role(request, project_id, allowed_roles=EDIT_ROLES)
    project = get_object_or_404(Project, id=project_id, workspace=request.pmo_workspace)
    payload = _json_body(request) or {}
    cal = get_calendar(project)

    parent = None
    if payload.get("parent_id"):
        parent = get_object_or_404(Task, id=payload["parent_id"], project=project)

    start = _parse_dt(payload.get("start_date"))
    end = _parse_dt(payload.get("end_date"), end_of_day=True)
    if start is None:
        base = parent.start_date if parent and parent.start_date else project.start_date
        base_date = local_date(base) if base else timezone.localtime(timezone.now()).date()
        start = to_start_dt(next_working_day(base_date, cal))
    else:
        start = to_start_dt(next_working_day(local_date(start), cal))
    if end is None or local_date(end) < local_date(start):
        end = to_end_dt(local_date(start))  # one-working-day task

    with transaction.atomic():
        after_id = payload.get("after_id")
        if after_id:
            after = get_object_or_404(Task, id=after_id, project=project)
            sort_order = after.sort_order + 1
            Task.objects.filter(project=project, sort_order__gte=sort_order).update(sort_order=F("sort_order") + 1)
        else:
            sort_order = (project.tasks.aggregate(m=Max("sort_order"))["m"] or 0) + 1

        sprint = None
        if payload.get("sprint_id"):
            sprint = get_object_or_404(Sprint, id=payload["sprint_id"], project=project)
        assignee = None
        if payload.get("assignee_id"):
            assignee = get_object_or_404(User, id=payload["assignee_id"])

        task = Task.objects.create(
            project=project,
            parent_task=parent,
            sprint=sprint,
            assignee=assignee,
            title=(payload.get("title") or "").strip() or "New Task",
            description=(payload.get("description") or "").strip() or None,
            status=(payload.get("status") or get_project_statuses(project)[0].key),
            start_date=start,
            end_date=end,
            sort_order=sort_order,
            is_milestone=bool(payload.get("is_milestone")),
        )
    if task.assignee is not None:
        notify_assignment(task, task.assignee, request.pmo_user)
    return JsonResponse({"ok": True, "task": _task_json(task, _predecessor_map(project))})


@require_POST
def task_update(request, project_id, task_id):
    """Single patch endpoint for inline grid edits: any subset of fields."""
    require_project_role(request, project_id, allowed_roles=EDIT_ROLES)
    project = get_object_or_404(Project, id=project_id, workspace=request.pmo_workspace)
    task = get_object_or_404(Task, id=task_id, project=project)
    payload = _json_body(request)
    if payload is None:
        return HttpResponseBadRequest("Expected a JSON body.")
    cal = get_calendar(project)

    dates_changed = False

    if "title" in payload and (payload["title"] or "").strip():
        task.title = payload["title"].strip()
    if "description" in payload:
        task.description = (payload["description"] or "").strip() or None
    if "status" in payload:
        smap = _status_category_map(project)
        if payload["status"] in smap:
            task.status = payload["status"]
            if smap[payload["status"]] == StatusCategory.DONE:
                task.progress = 100
    if "progress" in payload:
        try:
            task.progress = max(0, min(100, int(payload["progress"])))
        except (TypeError, ValueError):
            pass
    if "sprint_id" in payload:
        if payload["sprint_id"] is None:
            task.sprint = None
        else:
            task.sprint = get_object_or_404(Sprint, id=payload["sprint_id"], project=project)
    assignee_changed = False
    if "assignee_id" in payload:
        old_assignee_id = task.assignee_id
        if payload["assignee_id"] is None:
            task.assignee = None
        else:
            assignee = get_object_or_404(User, id=payload["assignee_id"])
            if not WorkspaceMember.objects.filter(workspace=request.pmo_workspace, user=assignee).exists():
                return HttpResponseBadRequest("That user isn't a member of this workspace.")
            task.assignee = assignee
        assignee_changed = task.assignee_id != old_assignee_id
        # keep the multi-assignment set in sync with the single primary
        task.assignments.all().delete()
        if task.assignee_id:
            TaskAssignment.objects.get_or_create(task=task, user_id=task.assignee_id)
    # Multiple resource assignments (MS Project style). Syncs the legacy single
    # `assignee` to the first for backward compatibility.
    if "assignee_ids" in payload:
        ids = payload["assignee_ids"] or []
        valid = list(WorkspaceMember.objects.filter(workspace=request.pmo_workspace, user_id__in=ids)
                     .values_list("user_id", flat=True))
        ordered = [i for i in ids if i in valid]
        task.assignments.exclude(user_id__in=ordered).delete()
        for uid in ordered:
            TaskAssignment.objects.get_or_create(task=task, user_id=uid)
        old_assignee_id = task.assignee_id
        task.assignee_id = ordered[0] if ordered else None
        assignee_changed = task.assignee_id != old_assignee_id
    if "is_milestone" in payload:
        task.is_milestone = bool(payload["is_milestone"])
        if task.is_milestone and task.start_date:
            task.end_date = task.start_date
            dates_changed = True
    constraint_changed = False
    if "constraint_type" in payload and payload["constraint_type"] in ConstraintType.values:
        task.constraint_type = payload["constraint_type"]
        if task.constraint_type in (ConstraintType.ASAP, ConstraintType.ALAP):
            task.constraint_date = None
        constraint_changed = True
    if "constraint_date" in payload:
        task.constraint_date = _parse_dt(payload["constraint_date"])
        constraint_changed = True
    if "deadline" in payload:
        task.deadline = _parse_dt(payload["deadline"], end_of_day=True)
    if "estimated_hours" in payload:
        try:
            task.estimated_hours = max(0.0, float(payload["estimated_hours"])) if payload["estimated_hours"] not in (None, "") else None
        except (TypeError, ValueError):
            pass
    if "story_points" in payload:
        try:
            task.story_points = max(0, int(payload["story_points"])) if payload["story_points"] not in (None, "") else None
        except (TypeError, ValueError):
            pass
    if "task_type" in payload and payload["task_type"] in TaskType.values:
        task.task_type = payload["task_type"]
    if "effort_driven" in payload:
        task.effort_driven = bool(payload["effort_driven"])
    if "scheduling_mode" in payload and payload["scheduling_mode"] in SchedulingMode.values:
        task.scheduling_mode = payload["scheduling_mode"]
    if "is_active" in payload:
        task.is_active = bool(payload["is_active"])
    if "fixed_cost" in payload:
        try:
            task.fixed_cost = max(0.0, float(payload["fixed_cost"] or 0))
        except (TypeError, ValueError):
            pass
    if "assignments" in payload and isinstance(payload["assignments"], list):
        # [{"user_id":.., "units":..}] — assign with per-resource units
        valid = set(WorkspaceMember.objects.filter(workspace=request.pmo_workspace).values_list("user_id", flat=True))
        seen = []
        for a in payload["assignments"]:
            try:
                uid = int(a.get("user_id"))
            except (TypeError, ValueError, AttributeError):
                continue
            if uid not in valid or uid in seen:
                continue
            try:
                u = max(0.0, float(a.get("units", 1.0)))
            except (TypeError, ValueError):
                u = 1.0
            TaskAssignment.objects.update_or_create(task=task, user_id=uid, defaults={"units": u})
            seen.append(uid)
        task.assignments.exclude(user_id__in=seen).delete()
        old_a = task.assignee_id
        task.assignee_id = seen[0] if seen else None
        assignee_changed = assignee_changed or (task.assignee_id != old_a)
    if "format" in payload and isinstance(payload["format"], dict):
        fmt = dict(task.format_json or {})
        for k in ("bold", "italic"):
            if k in payload["format"]:
                fmt[k] = bool(payload["format"][k])
        for k in ("color", "bg", "family", "valign"):
            if k in payload["format"]:
                fmt[k] = payload["format"][k] or None
        if "size" in payload["format"]:
            try:
                fmt["size"] = max(8, min(28, int(payload["format"]["size"]))) if payload["format"]["size"] else None
            except (TypeError, ValueError):
                pass
        task.format_json = {k: v for k, v in fmt.items() if v} or None

    if "start_date" in payload:
        new_start = _parse_dt(payload["start_date"])
        if new_start:
            new_start_date = next_working_day(local_date(new_start), cal)
            if task.start_date and task.end_date and "end_date" not in payload and "duration" not in payload:
                # Moving the start keeps the working-day duration (MS Project behaviour).
                duration = count_working_days(local_date(task.start_date), local_date(task.end_date), cal)
                task.end_date = to_end_dt(add_working_days(new_start_date, duration, cal))
            task.start_date = to_start_dt(new_start_date)
            dates_changed = True
    if "end_date" in payload:
        new_end = _parse_dt(payload["end_date"], end_of_day=True)
        if new_end:
            task.end_date = to_end_dt(local_date(new_end))
            dates_changed = True
    if "duration" in payload and task.start_date:
        try:
            duration = max(1, int(payload["duration"]))
            task.end_date = to_end_dt(add_working_days(local_date(task.start_date), duration, cal))
            dates_changed = True
        except (TypeError, ValueError):
            pass

    if task.start_date and task.end_date and task.end_date < task.start_date:
        return HttpResponseBadRequest("Finish date can't be before the start date.")

    if "parent_id" in payload:
        if payload["parent_id"] is None:
            task.parent_task = None
        else:
            parent = get_object_or_404(Task, id=payload["parent_id"], project=project)
            node = parent
            while node is not None:
                if node.id == task.id:
                    return HttpResponseBadRequest("A task can't be nested inside itself.")
                node = node.parent_task
            task.parent_task = parent

    task.save()

    if assignee_changed and task.assignee is not None:
        notify_assignment(task, task.assignee, request.pmo_user)

    # A new/changed constraint moves the task itself to where the constraint wants it.
    if constraint_changed and task.start_date and task.end_date:
        required = required_start_for(task, cal, {})
        if required and required != local_date(task.start_date):
            duration = count_working_days(local_date(task.start_date), local_date(task.end_date), cal)
            task.start_date = to_start_dt(required)
            task.end_date = to_end_dt(add_working_days(required, duration, cal))
            task.save()
            dates_changed = True

    if dates_changed and task.start_date and task.end_date:
        reschedule_task_and_successors(task.id, task.start_date, task.end_date)

    # Replace the predecessor list if one was sent: [{"id": 3, "type": "FS"}, ...]
    if "predecessors" in payload:
        long = {"FS": DependencyType.FINISH_TO_START, "SS": DependencyType.START_TO_START,
                "FF": DependencyType.FINISH_TO_FINISH, "SF": DependencyType.START_TO_FINISH}
        wanted = []
        for p in payload["predecessors"] or []:
            pred_id = p.get("id") if isinstance(p, dict) else p
            dep_type = long.get((p.get("type") or "FS").upper() if isinstance(p, dict) else "FS",
                                DependencyType.FINISH_TO_START)
            try:
                lag = int(p.get("lag") or 0) if isinstance(p, dict) else 0
            except (TypeError, ValueError):
                lag = 0
            if not Task.objects.filter(id=pred_id, project=project).exists():
                return HttpResponseBadRequest(f"No task #{pred_id} in this project.")
            if _would_create_cycle(project.id, pred_id, task.id):
                return HttpResponseBadRequest("That predecessor would create a circular dependency.")
            wanted.append((pred_id, dep_type, lag))
        with transaction.atomic():
            TaskDependency.objects.filter(to_task=task).delete()
            for pred_id, dep_type, lag in wanted:
                TaskDependency.objects.create(from_task_id=pred_id, to_task=task, type=dep_type, lag_days=lag)
        # Enforce the new constraints immediately (push this task if needed).
        for pred_id, _t, _l in wanted:
            pred = Task.objects.get(id=pred_id)
            if pred.start_date and pred.end_date:
                reschedule_task_and_successors(pred.id, pred.start_date, pred.end_date)

    # Custom field values: {"custom_values": {"<field_id>": "value", ...}}
    if "custom_values" in payload and isinstance(payload["custom_values"], dict):
        for fid, value in payload["custom_values"].items():
            field = CustomField.objects.filter(id=fid, project=project).first()
            if field is None:
                continue
            if value in (None, ""):
                TaskCustomFieldValue.objects.filter(task=task, custom_field=field).delete()
            else:
                TaskCustomFieldValue.objects.update_or_create(
                    task=task, custom_field=field, defaults={"value": str(value)}
                )

    task.refresh_from_db()
    custom = {task.id: {v.custom_field_id: v.value for v in task.custom_values.all()}}
    logged = {task.id: task.time_entries.aggregate(t=Sum("minutes"))["t"] or 0}
    return JsonResponse({"ok": True, "task": _task_json(task, _predecessor_map(project),
                                                        {"custom": custom, "logged": logged})})


@require_POST
def task_delete(request, project_id, task_id):
    require_project_role(request, project_id, allowed_roles=EDIT_ROLES)
    task = get_object_or_404(Task, id=task_id, project_id=project_id, project__workspace=request.pmo_workspace)
    task.delete()  # cascades to subtasks + dependencies
    return JsonResponse({"ok": True})


@require_POST
def task_reschedule(request, project_id, task_id):
    """Called when a bar is dragged/resized on the Gantt. Cascades to dependents."""
    require_project_role(request, project_id, allowed_roles=EDIT_ROLES)
    task = get_object_or_404(Task, id=task_id, project_id=project_id)
    payload = _json_body(request)
    if payload is None:
        return HttpResponseBadRequest("Expected JSON body with start_date/end_date.")
    next_start = _parse_dt(payload.get("start_date"))
    next_end = _parse_dt(payload.get("end_date"), end_of_day=True)
    if not next_start or not next_end:
        return HttpResponseBadRequest("Invalid date format.")

    reschedule_task_and_successors(task.id, next_start, next_end)
    return JsonResponse({"ok": True})


@require_POST
def baseline_set(request, project_id):
    """Snapshot every task's current dates as the baseline (like MS Project's Set Baseline)."""
    require_project_role(request, project_id, allowed_roles=EDIT_ROLES)
    project = get_object_or_404(Project, id=project_id, workspace=request.pmo_workspace)
    n = 0
    with transaction.atomic():
        for t in project.tasks.all():
            t.baseline_start = t.start_date
            t.baseline_end = t.end_date
            t.save(update_fields=["baseline_start", "baseline_end"])
            n += 1
    return JsonResponse({"ok": True, "tasks": n})


@require_POST
def task_log_time(request, project_id, task_id):
    require_project_role(request, project_id, allowed_roles=EDIT_ROLES + ["VIEWER", "COMMENTOR"])
    task = get_object_or_404(Task, id=task_id, project_id=project_id, project__workspace=request.pmo_workspace)
    payload = _json_body(request)
    if payload is None:
        return HttpResponseBadRequest("Expected a JSON body.")
    try:
        minutes = int(payload.get("minutes"))
        if minutes <= 0:
            raise ValueError
    except (TypeError, ValueError):
        return HttpResponseBadRequest("minutes must be a positive integer.")
    TimeEntry.objects.create(
        project=task.project, task=task, user=request.pmo_user,
        entry_date=_parse_dt(payload.get("date")) or timezone.now(),
        minutes=minutes, note=(payload.get("note") or "").strip() or None,
    )
    total = task.time_entries.aggregate(t=Sum("minutes"))["t"] or 0
    return JsonResponse({"ok": True, "logged_minutes": total})


@require_POST
def task_comment_add(request, project_id, task_id):
    """Any project member (any role) can comment on a task (change request #12)."""
    project = get_object_or_404(Project, id=project_id, workspace=request.pmo_workspace)
    _require_project_access(request, project)
    task = get_object_or_404(Task, id=task_id, project=project)
    payload = _json_body(request)
    if payload is None or not (payload.get("body") or "").strip():
        return HttpResponseBadRequest("Comment text is required.")
    c = TaskComment.objects.create(task=task, user=request.pmo_user, body=payload["body"].strip())
    return JsonResponse({"ok": True, "id": c.id})


@require_POST
def task_attachment_add(request, project_id, task_id):
    """Link-based attachment (name + URL). Binary uploads go to Azure Blob in prod."""
    require_project_role(request, project_id, allowed_roles=EDIT_ROLES)
    task = get_object_or_404(Task, id=task_id, project_id=project_id,
                             project__workspace=request.pmo_workspace)
    payload = _json_body(request)
    if payload is None or not (payload.get("name") or "").strip() or not (payload.get("url") or "").strip():
        return HttpResponseBadRequest("Attachment needs a name and a URL.")
    a = TaskAttachment.objects.create(task=task, file_name=payload["name"].strip(), blob_url=payload["url"].strip())
    return JsonResponse({"ok": True, "id": a.id})


@require_POST
def task_attachment_delete(request, project_id, task_id, attachment_id):
    require_project_role(request, project_id, allowed_roles=EDIT_ROLES)
    a = get_object_or_404(TaskAttachment, id=attachment_id, task_id=task_id,
                          task__project_id=project_id, task__project__workspace=request.pmo_workspace)
    a.delete()
    return JsonResponse({"ok": True})


@require_POST
def field_create(request, project_id):
    require_project_role(request, project_id, allowed_roles=EDIT_ROLES)
    project = get_object_or_404(Project, id=project_id, workspace=request.pmo_workspace)
    payload = _json_body(request)
    if payload is None or not (payload.get("name") or "").strip():
        return HttpResponseBadRequest("A field name is required.")
    ftype = payload.get("type") if payload.get("type") in CustomFieldType.values else CustomFieldType.TEXT
    options = None
    if ftype == CustomFieldType.SELECT:
        options = [str(o).strip() for o in (payload.get("options") or []) if str(o).strip()]
        if not options:
            return HttpResponseBadRequest("A choice field needs at least one option.")
    field = CustomField.objects.create(
        project=project, name=payload["name"].strip(), type=ftype, options_json=options)
    return JsonResponse({"ok": True, "id": field.id})


@require_POST
def field_delete(request, project_id, field_id):
    require_project_role(request, project_id, allowed_roles=EDIT_ROLES)
    field = get_object_or_404(CustomField, id=field_id, project_id=project_id,
                              project__workspace=request.pmo_workspace)
    field.delete()
    return JsonResponse({"ok": True})


@ensure_csrf_cookie
def project_network(request, project_id):
    """Network (PERT) diagram of the task dependency graph."""
    project = get_object_or_404(Project, id=project_id, workspace=request.pmo_workspace)
    _require_project_access(request, project)
    link = PortfolioProject.objects.filter(project=project).select_related("portfolio").first()
    return render(request, "pmo/project_network.html", {
        "project": project,
        "portfolio": link.portfolio if link else None,
    })


# ---------------------------------------------------------------------------
# Excel import / export (change request: MS Project-style Excel round-trip)
# ---------------------------------------------------------------------------

IMPORT_HEADER_ALIASES = {
    "name": {"task name", "name", "title", "task"},
    "wbs": {"wbs", "outline number"},
    "outline": {"outline level", "level", "indent"},
    "duration": {"duration", "dur", "duration (days)", "days"},
    "start": {"start", "start date", "start_date", "begin"},
    "end": {"finish", "finish date", "end", "end date", "end_date", "due"},
    "pred": {"predecessors", "predecessor", "pred", "depends on"},
    "resource": {"resource", "resource names", "assignee", "owner", "assigned to"},
    "progress": {"% complete", "percent complete", "progress", "%", "% done"},
    "description": {"description", "notes", "details"},
    "milestone": {"milestone", "is milestone"},
}


@require_POST
def excel_import(request, project_id):
    """Import an Excel task list (like MS Project's Excel import wizard).

    Recognised columns (any order, case-insensitive): Task Name, WBS or
    Outline Level, Duration, Start, Finish, Predecessors, Resource, % Complete,
    Description, Milestone. Hierarchy comes from a WBS column (1.2.3) or an
    Outline Level column; otherwise the list imports flat."""
    require_project_role(request, project_id, allowed_roles=EDIT_ROLES)
    project = get_object_or_404(Project, id=project_id, workspace=request.pmo_workspace)
    if "file" not in request.FILES:
        return HttpResponseBadRequest("Upload an .xlsx file in the 'file' field.")
    try:
        from openpyxl import load_workbook
    except ImportError:
        return HttpResponseBadRequest(
            "Excel support isn't installed on the server — run: pip install openpyxl"
        )
    try:
        wb = load_workbook(request.FILES["file"], data_only=True)
    except Exception:
        return HttpResponseBadRequest("Couldn't read that file — it must be a valid .xlsx workbook.")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return HttpResponseBadRequest("The sheet is empty.")

    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    cols = {}
    for key, aliases in IMPORT_HEADER_ALIASES.items():
        for i, h in enumerate(header):
            if h in aliases:
                cols[key] = i
                break
    if "name" not in cols:
        return HttpResponseBadRequest("Couldn't find a 'Task Name' column in the first row.")

    def cell(row, key):
        i = cols.get(key)
        return row[i] if i is not None and i < len(row) else None

    cal = get_calendar(project)
    sort_order = (project.tasks.aggregate(m=Max("sort_order"))["m"] or 0)
    member_by_name = {}
    for m in WorkspaceMember.objects.filter(workspace=request.pmo_workspace).select_related("user"):
        member_by_name[m.user.name.strip().lower()] = m.user
        member_by_name[m.user.email.strip().lower()] = m.user

    def as_date(value, end=False):
        if value is None or value == "":
            return None
        if isinstance(value, datetime):
            d = value.date()
        else:
            parsed = _parse_dt(str(value).strip()[:10])
            if parsed is None:
                return None
            d = local_date(parsed)
        return to_end_dt(d) if end else to_start_dt(d)

    created, warnings = [], []
    parent_stack = []  # [(depth, task)]
    wbs_map = {}       # wbs string or row number -> task
    pending_preds = []

    with transaction.atomic():
        # Any header column that isn't a known field becomes a custom column,
        # imported with its name and values as-is (great for a new project).
        orig_header = [str(c).strip() if c is not None else "" for c in rows[0]]
        known_idx = set(cols.values())
        existing_fields = {f.name.strip().lower(): f for f in project.custom_fields.all()}
        extra_cols = []  # [(col_index, CustomField)]
        for i, hname in enumerate(orig_header):
            if i in known_idx or not hname:
                continue
            f = existing_fields.get(hname.strip().lower())
            if f is None:
                f = CustomField.objects.create(project=project, name=hname[:255], type=CustomFieldType.TEXT)
                existing_fields[hname.strip().lower()] = f
            extra_cols.append((i, f))

        for idx, row in enumerate(rows[1:], start=1):
            name = cell(row, "name")
            if name is None or not str(name).strip():
                continue
            wbs_raw = str(cell(row, "wbs") or "").strip().rstrip(".")
            if wbs_raw:
                depth = wbs_raw.count(".")
            else:
                try:
                    depth = max(0, int(cell(row, "outline") or 1) - 1)
                except (TypeError, ValueError):
                    depth = 0
            while parent_stack and parent_stack[-1][0] >= depth:
                parent_stack.pop()
            parent = parent_stack[-1][1] if parent_stack else None

            start = as_date(cell(row, "start"))
            end = as_date(cell(row, "end"), end=True)
            duration = None
            try:
                duration = int(float(str(cell(row, "duration")).replace("d", "").strip()))
            except (TypeError, ValueError):
                pass
            if start is None:
                base = parent.start_date if parent and parent.start_date else project.start_date
                start = to_start_dt(next_working_day(local_date(base) if base else timezone.localtime(timezone.now()).date(), cal))
            if end is None:
                end = to_end_dt(add_working_days(local_date(start), duration or 1, cal))

            progress = 0
            try:
                p = cell(row, "progress")
                if p is not None:
                    p = float(p)
                    progress = int(p * 100) if p <= 1 else int(p)
            except (TypeError, ValueError):
                pass

            assignee = None
            res = cell(row, "resource")
            if res:
                assignee = member_by_name.get(str(res).strip().lower())
                if assignee is None:
                    warnings.append(f"Row {idx}: resource '{res}' isn't a workspace member — left unassigned.")

            milestone = str(cell(row, "milestone") or "").strip().lower() in ("yes", "y", "true", "1")

            sort_order += 1
            task = Task.objects.create(
                project=project, parent_task=parent,
                title=str(name).strip()[:500],
                description=(str(cell(row, "description") or "").strip() or None),
                start_date=start, end_date=start if milestone else end,
                progress=max(0, min(100, progress)),
                sort_order=sort_order, is_milestone=milestone, assignee=assignee,
            )
            created.append(task)
            for ci, f in extra_cols:
                val = row[ci] if ci < len(row) else None
                if val is not None and str(val).strip():
                    TaskCustomFieldValue.objects.create(task=task, custom_field=f, value=str(val).strip())
            wbs_map[wbs_raw or str(idx)] = task
            wbs_map[str(idx)] = task
            parent_stack.append((depth, task))
            pred_raw = str(cell(row, "pred") or "").strip()
            if pred_raw:
                pending_preds.append((task, pred_raw, idx))

        # Resolve predecessors after all rows exist (WBS refs or row numbers).
        long = {"FS": DependencyType.FINISH_TO_START, "SS": DependencyType.START_TO_START,
                "FF": DependencyType.FINISH_TO_FINISH, "SF": DependencyType.START_TO_FINISH}
        import re
        for task, raw, idx in pending_preds:
            for piece in re.split(r"[,;]", raw):
                m = re.match(r"^([\d.]+)\s*(FS|SS|FF|SF)?\s*([+-]\d+)?d?$", piece.strip(), re.I)
                if not m:
                    warnings.append(f"Row {idx}: couldn't read predecessor '{piece.strip()}'.")
                    continue
                pred = wbs_map.get(m.group(1))
                if pred is None or pred.id == task.id:
                    warnings.append(f"Row {idx}: predecessor '{m.group(1)}' not found.")
                    continue
                TaskDependency.objects.create(
                    from_task=pred, to_task=task,
                    type=long.get((m.group(2) or "FS").upper(), DependencyType.FINISH_TO_START),
                    lag_days=int(m.group(3)) if m.group(3) else 0,
                )

    # Let the scheduler enforce the imported dependencies.
    for task, _raw, _idx in pending_preds:
        for dep in TaskDependency.objects.filter(to_task=task).select_related("from_task"):
            if dep.from_task.start_date and dep.from_task.end_date:
                reschedule_task_and_successors(dep.from_task.id, dep.from_task.start_date, dep.from_task.end_date)
                break
    return JsonResponse({"ok": True, "created": len(created), "warnings": warnings[:20]})


def excel_export(request, project_id):
    """Download the task list as .xlsx (like MS Project's Save As Excel).
    Falls back to CSV when openpyxl isn't installed, so export always works."""
    from django.http import HttpResponse
    project = get_object_or_404(Project, id=project_id, workspace=request.pmo_workspace)
    _require_project_access(request, project)

    cal = get_calendar(project)
    predecessors = _predecessor_map(project)
    tasks = list(project.tasks.select_related("assignee"))
    children = {}
    for t in tasks:
        children.setdefault(t.parent_task_id, []).append(t)
    for v in children.values():
        v.sort(key=lambda t: (t.sort_order, t.id))

    flat = []
    def walk(parent_id, depth, prefix):
        for i, t in enumerate(children.get(parent_id, []), start=1):
            wbs = f"{prefix}.{i}" if prefix else str(i)
            flat.append((t, depth, wbs))
            walk(t.id, depth + 1, wbs)
    walk(None, 0, "")
    wbs_by_id = {t.id: w for t, _d, w in flat}
    headers = ["WBS", "Task Name", "Outline Level", "Duration", "Start", "Finish",
               "Predecessors", "Resource", "% Complete", "Status", "Milestone",
               "Work (h)", "Points", "Deadline", "Description"]
    safe = "".join(ch for ch in project.name if ch.isalnum() or ch in " -_")[:60] or "project"
    status_names = {s.key: s.name for s in get_project_statuses(project)}

    def data_rows():
        for t, depth, wbs in flat:
            s, e = local_date(t.start_date), local_date(t.end_date)
            preds = ",".join(
                f"{wbs_by_id.get(p['id'], '?')}{'' if p['type'] == 'FS' and not p['lag'] else p['type']}"
                f"{('+' + str(p['lag'])) if p['lag'] > 0 else (str(p['lag']) if p['lag'] < 0 else '')}"
                for p in predecessors.get(t.id, [])
            )
            yield [
                wbs, ("    " * depth) + t.title, depth + 1,
                0 if t.is_milestone else (count_working_days(s, e, cal) if s and e else None),
                s, e, preds,
                t.assignee.name if t.assignee else "",
                t.progress, status_names.get(t.status, t.status), "Yes" if t.is_milestone else "",
                t.estimated_hours, t.story_points,
                local_date(t.deadline), t.description or "",
            ]

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font as XFont, PatternFill
    except ImportError:
        # CSV fallback keeps export working on any environment.
        import csv
        from io import StringIO
        sio = StringIO()
        w = csv.writer(sio)
        w.writerow(headers)
        for row in data_rows():
            w.writerow(["" if v is None else v for v in row])
        resp = HttpResponse(sio.getvalue(), content_type="text/csv")
        resp["Content-Disposition"] = f'attachment; filename="{safe} - task list.csv"'
        return resp

    wb = Workbook()
    ws = wb.active
    ws.title = "Task List"
    ws.append(headers)
    hfill = PatternFill("solid", fgColor="1D4ED8")
    for c in ws[1]:
        c.font = XFont(bold=True, color="FFFFFF")
        c.fill = hfill
    for row in data_rows():
        ws.append(row)
    widths = [8, 42, 7, 9, 12, 12, 14, 18, 8, 12, 9, 8, 7, 12, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    resp = HttpResponse(buf.getvalue(),
                        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{safe} - task list.xlsx"'
    return resp


@require_POST
def task_reorder(request, project_id):
    """[{"id": 1, "sort_order": 0, "parent_id": null}, ...] — from drag & drop."""
    require_project_role(request, project_id, allowed_roles=EDIT_ROLES)
    payload = _json_body(request)
    if payload is None or "order" not in payload:
        return HttpResponseBadRequest('Expected JSON body: {"order": [{"id":.., "sort_order":..}]}')

    with transaction.atomic():
        for item in payload["order"]:
            fields = {"sort_order": item["sort_order"]}
            if "parent_id" in item:
                fields["parent_task_id"] = item["parent_id"]
            Task.objects.filter(id=item["id"], project_id=project_id).update(**fields)
    return JsonResponse({"ok": True})


# ---------------------------------------------------------------------------
# Real-time collaboration (rTc) — presence heartbeat + change revision
# ---------------------------------------------------------------------------

PRESENCE_TTL = 20  # seconds a heartbeat is considered "live"


def _plan_revision(project):
    """A cheap fingerprint of the plan that changes on any task edit/add/delete,
    so live clients know when to refetch without a websocket."""
    agg = project.tasks.aggregate(m=Max("updated_at"), n=Sum("id"))
    latest = agg["m"].isoformat() if agg["m"] else "0"
    count = project.tasks.count()
    return f"{latest}:{count}:{agg['n'] or 0}"


@require_POST
def presence(request, project_id):
    """Heartbeat for real-time collaboration. Records that the current user is
    viewing the plan and returns everyone active in the last few seconds plus the
    current plan revision. Backed by the cache — no websocket layer required."""
    from django.core.cache import cache
    import time as _time

    project = get_object_or_404(Project, id=project_id, workspace=request.pmo_workspace)
    _require_project_access(request, project)

    body = _json_body(request) or {}
    leaving = bool(body.get("leaving"))
    editing = bool(body.get("editing"))
    me = request.pmo_user

    key = f"pmo-presence-{project_id}"
    now = _time.time()
    roster = cache.get(key) or {}
    # prune stale entries
    roster = {uid: info for uid, info in roster.items() if now - info.get("ts", 0) < PRESENCE_TTL}
    if leaving:
        roster.pop(me.id, None)
    else:
        roster[me.id] = {"name": me.name, "ts": now, "editing": editing}
    cache.set(key, roster, PRESENCE_TTL * 3)

    collaborators = [
        {"id": uid, "name": info["name"], "editing": info.get("editing", False), "you": uid == me.id}
        for uid, info in sorted(roster.items(), key=lambda kv: kv[1]["name"].lower())
    ]
    return JsonResponse({"ok": True, "collaborators": collaborators, "revision": _plan_revision(project)})


# ---------------------------------------------------------------------------
# MS Project compatibility — MSPDI (.xml) export / import
# Microsoft Project opens/saves this XML natively (and can Save-As .mpp),
# which is the practical, cross-version-safe route to "MPP compatible".
# ---------------------------------------------------------------------------

_MSP_NS = "http://schemas.microsoft.com/project"
# MSPDI PredecessorLink Type codes
_MSP_LINK_CODE = {"FF": 0, "FS": 1, "SF": 2, "SS": 3}
_MSP_LINK_TYPE = {0: "FF", 1: "FS", 2: "SF", 3: "SS"}


def _msp_dt(d, hour=8):
    if not d:
        return None
    return f"{d.isoformat()}T{hour:02d}:00:00"


def msp_export(request, project_id):
    """Export the plan as MS Project MSPDI XML (opens natively in Microsoft Project)."""
    from django.http import HttpResponse
    from xml.sax.saxutils import escape

    project = get_object_or_404(Project, id=project_id, workspace=request.pmo_workspace)
    _require_project_access(request, project)

    cal = get_calendar(project)
    predecessors = _predecessor_map(project)
    tasks = list(project.tasks.select_related("assignee"))
    children = {}
    for t in tasks:
        children.setdefault(t.parent_task_id, []).append(t)
    for v in children.values():
        v.sort(key=lambda t: (t.sort_order, t.id))

    flat = []
    def walk(parent_id, depth, prefix):
        for i, t in enumerate(children.get(parent_id, []), start=1):
            wbs = f"{prefix}.{i}" if prefix else str(i)
            flat.append((t, depth, wbs))
            walk(t.id, depth + 1, wbs)
    walk(None, 0, "")

    is_summary = {t.id for t in tasks if children.get(t.id)}
    uid_to_seq = {t.id: i for i, (t, _d, _w) in enumerate(flat, start=1)}

    out = ['<?xml version="1.0" encoding="UTF-8" standalone="yes"?>']
    out.append(f'<Project xmlns="{_MSP_NS}">')
    out.append("<SaveVersion>14</SaveVersion>")
    out.append(f"<Name>{escape(project.name)}</Name>")
    out.append(f"<Title>{escape(project.name)}</Title>")
    out.append("<ScheduleFromStart>1</ScheduleFromStart>")
    if project.start_date:
        out.append(f"<StartDate>{_msp_dt(local_date(project.start_date))}</StartDate>")
    out.append("<CalendarUID>1</CalendarUID>")
    out.append("<DurationFormat>7</DurationFormat>")   # 7 = days

    # --- Minimal Standard calendar so Project doesn't prompt ---
    working = set(cal[0]) if isinstance(cal, tuple) and cal[0] else {1, 2, 3, 4, 5}
    out.append("<Calendars><Calendar><UID>1</UID><Name>Standard</Name><IsBaseCalendar>1</IsBaseCalendar>")
    out.append("<WeekDays>")
    for dow in range(1, 8):  # MSPDI DayType: 1=Sun..7=Sat
        iso_dow = 7 if dow == 1 else dow - 1  # map to ISO 1=Mon..7=Sun
        is_work = iso_dow in working
        out.append(f"<WeekDay><DayType>{dow}</DayType><DayWorking>{1 if is_work else 0}</DayWorking>")
        if is_work:
            out.append("<WorkingTimes>"
                       "<WorkingTime><FromTime>08:00:00</FromTime><ToTime>12:00:00</ToTime></WorkingTime>"
                       "<WorkingTime><FromTime>13:00:00</FromTime><ToTime>17:00:00</ToTime></WorkingTime>"
                       "</WorkingTimes>")
        out.append("</WeekDay>")
    out.append("</WeekDays></Calendar></Calendars>")

    # --- Tasks ---
    out.append("<Tasks>")
    for t, depth, wbs in flat:
        s, e = local_date(t.start_date), local_date(t.end_date)
        summary = t.id in is_summary
        days = 0 if t.is_milestone else (count_working_days(s, e, cal) if s and e else 1)
        hours = int((days or 0) * 8)
        out.append("<Task>")
        out.append(f"<UID>{t.id}</UID>")
        out.append(f"<ID>{uid_to_seq[t.id]}</ID>")
        out.append(f"<Name>{escape(t.title)}</Name>")
        out.append("<Active>1</Active>")
        out.append("<Manual>0</Manual>")
        out.append(f"<OutlineLevel>{depth + 1}</OutlineLevel>")
        out.append(f"<OutlineNumber>{wbs}</OutlineNumber>")
        out.append(f"<WBS>{wbs}</WBS>")
        out.append(f"<Summary>{1 if summary else 0}</Summary>")
        out.append(f"<Milestone>{1 if t.is_milestone else 0}</Milestone>")
        out.append(f"<PercentComplete>{int(t.progress or 0)}</PercentComplete>")
        if s:
            out.append(f"<Start>{_msp_dt(s, 8)}</Start>")
        if e:
            out.append(f"<Finish>{_msp_dt(e, 17)}</Finish>")
        out.append(f"<Duration>PT{hours}H0M0S</Duration>")
        out.append("<DurationFormat>7</DurationFormat>")
        if t.deadline:
            out.append(f"<Deadline>{_msp_dt(local_date(t.deadline), 17)}</Deadline>")
        if t.description:
            out.append(f"<Notes>{escape(t.description)}</Notes>")
        for p in predecessors.get(t.id, []):
            if p["id"] in uid_to_seq:
                out.append("<PredecessorLink>"
                           f"<PredecessorUID>{p['id']}</PredecessorUID>"
                           f"<Type>{_MSP_LINK_CODE.get(p['type'], 1)}</Type>"
                           f"<LinkLag>{int(p.get('lag', 0)) * 4800}</LinkLag>"
                           "<LagFormat>7</LagFormat>"
                           "</PredecessorLink>")
        out.append("</Task>")
    out.append("</Tasks>")

    # --- Resources + Assignments ---
    resources = {}
    for t in tasks:
        if t.assignee_id and t.assignee_id not in resources:
            resources[t.assignee_id] = t.assignee.name
    if resources:
        out.append("<Resources>")
        for rid, name in resources.items():
            out.append(f"<Resource><UID>{rid}</UID><ID>{rid}</ID><Name>{escape(name)}</Name><Type>1</Type></Resource>")
        out.append("</Resources>")
        out.append("<Assignments>")
        aid = 1
        for t in tasks:
            if t.assignee_id:
                out.append(f"<Assignment><UID>{aid}</UID><TaskUID>{t.id}</TaskUID>"
                           f"<ResourceUID>{t.assignee_id}</ResourceUID><Units>1</Units></Assignment>")
                aid += 1
        out.append("</Assignments>")

    out.append("</Project>")

    safe = "".join(ch for ch in project.name if ch.isalnum() or ch in " -_")[:60] or "project"
    resp = HttpResponse("".join(out), content_type="application/xml")
    resp["Content-Disposition"] = f'attachment; filename="{safe}.xml"'
    return resp


@require_POST
def msp_import(request, project_id):
    """Import tasks from an MS Project MSPDI (.xml) export — outline, dates,
    % complete, milestones and predecessor links."""
    import xml.etree.ElementTree as ET

    require_project_role(request, project_id, allowed_roles=EDIT_ROLES)
    project = get_object_or_404(Project, id=project_id, workspace=request.pmo_workspace)
    if "file" not in request.FILES:
        return HttpResponseBadRequest("Upload an MS Project .xml (MSPDI) file in the 'file' field.")
    try:
        tree = ET.parse(request.FILES["file"])
    except Exception:
        return HttpResponseBadRequest("Couldn't parse that file — export it from MS Project as XML (MSPDI).")
    root = tree.getroot()
    ns = {"m": _MSP_NS}

    def txt(node, tag, default=None):
        el = node.find(f"m:{tag}", ns)
        return el.text if el is not None and el.text is not None else default

    tasks_el = root.find("m:Tasks", ns)
    if tasks_el is None:
        return HttpResponseBadRequest("No <Tasks> section found — is this an MSPDI file?")

    cal = get_calendar(project)
    sort_order = (project.tasks.aggregate(m=Max("sort_order"))["m"] or 0)

    def as_date(value, end=False):
        if not value:
            return None
        parsed = _parse_dt(str(value).strip()[:10])
        if parsed is None:
            return None
        d = local_date(parsed)
        return to_end_dt(d) if end else to_start_dt(d)

    created, warnings = [], []
    parent_stack = []          # [(outline_level, task)]
    uid_to_task = {}
    pending_preds = []         # (task, [(pred_uid, type_code, lag), ...])

    with transaction.atomic():
        for tnode in tasks_el.findall("m:Task", ns):
            name = txt(tnode, "Name")
            if not name or not name.strip():
                continue
            try:
                level = max(1, int(txt(tnode, "OutlineLevel", "1")))
            except (TypeError, ValueError):
                level = 1
            # UID 0 with no name is the project summary in some files — skip empties above
            while parent_stack and parent_stack[-1][0] >= level:
                parent_stack.pop()
            parent = parent_stack[-1][1] if parent_stack else None

            start = as_date(txt(tnode, "Start"))
            end = as_date(txt(tnode, "Finish"), end=True)
            if start is None:
                base = parent.start_date if parent and parent.start_date else project.start_date
                start = to_start_dt(next_working_day(local_date(base) if base else timezone.localtime(timezone.now()).date(), cal))
            if end is None or end < start:
                end = to_end_dt(add_working_days(local_date(start), 1, cal))

            milestone = (txt(tnode, "Milestone", "0") == "1")
            try:
                progress = int(float(txt(tnode, "PercentComplete", "0")))
            except (TypeError, ValueError):
                progress = 0

            sort_order += 1
            task = Task.objects.create(
                project=project, parent_task=parent,
                title=name.strip()[:500],
                description=(txt(tnode, "Notes") or None),
                start_date=start, end_date=start if milestone else end,
                progress=max(0, min(100, progress)),
                sort_order=sort_order, is_milestone=milestone,
            )
            created.append(task)
            uid = txt(tnode, "UID")
            if uid is not None:
                uid_to_task[uid] = task
            parent_stack.append((level, task))

            preds = []
            for link in tnode.findall("m:PredecessorLink", ns):
                puid = txt(link, "PredecessorUID")
                try:
                    code = int(txt(link, "Type", "1"))
                except (TypeError, ValueError):
                    code = 1
                try:
                    lag = int(int(txt(link, "LinkLag", "0")) / 4800)
                except (TypeError, ValueError):
                    lag = 0
                if puid is not None:
                    preds.append((puid, code, lag))
            if preds:
                pending_preds.append((task, preds))

        long = {"FS": DependencyType.FINISH_TO_START, "SS": DependencyType.START_TO_START,
                "FF": DependencyType.FINISH_TO_FINISH, "SF": DependencyType.START_TO_FINISH}
        for task, preds in pending_preds:
            for puid, code, lag in preds:
                pred = uid_to_task.get(puid)
                if pred is None or pred.id == task.id:
                    warnings.append(f"'{task.title}': predecessor UID {puid} not found in file.")
                    continue
                TaskDependency.objects.create(
                    from_task=pred, to_task=task,
                    type=long.get(_MSP_LINK_TYPE.get(code, "FS"), DependencyType.FINISH_TO_START),
                    lag_days=lag,
                )

    for task, preds in pending_preds:
        for dep in TaskDependency.objects.filter(to_task=task).select_related("from_task"):
            if dep.from_task.start_date and dep.from_task.end_date:
                reschedule_task_and_successors(dep.from_task.id, dep.from_task.start_date, dep.from_task.end_date)
                break

    return JsonResponse({"ok": True, "created": len(created), "warnings": warnings[:20]})


# ---------------------------------------------------------------------------
# Custom task statuses — bulk save (create / rename / recolour / reorder / delete)
# ---------------------------------------------------------------------------

@require_POST
def statuses_save(request, project_id):
    """Replace this project's status set with the supplied ordered list.
    Items with a known 'key' are updated in place; new items get a generated key;
    omitted statuses are deleted and their tasks reassigned to a safe fallback."""
    require_project_role(request, project_id, allowed_roles=EDIT_ROLES)
    project = get_object_or_404(Project, id=project_id, workspace=request.pmo_workspace)
    payload = _json_body(request) or {}
    incoming = payload.get("statuses")
    if not isinstance(incoming, list) or not incoming:
        return HttpResponseBadRequest("Provide a non-empty 'statuses' list.")

    existing = {s.key: s for s in get_project_statuses(project)}
    valid_cats = set(StatusCategory.values)
    taken = set(existing.keys())
    kept_keys = set()

    with transaction.atomic():
        order = 0
        for item in incoming:
            name = str(item.get("name") or "").strip()[:60]
            if not name:
                continue
            cat = item.get("category")
            if cat not in valid_cats:
                cat = StatusCategory.NOT_STARTED
            color = str(item.get("color") or "#94a3b8")[:9]
            key = item.get("key")
            if key and key in existing:
                s = existing[key]
                s.name, s.category, s.color, s.sort_order = name, cat, color, order
                s.save()
            else:
                key = _slug_status_key(name, taken)
                taken.add(key)
                WorkflowStatus.objects.create(
                    project=project, key=key, name=name, category=cat, color=color, sort_order=order)
            kept_keys.add(key)
            order += 1

        if not kept_keys:
            return HttpResponseBadRequest("At least one status is required.")

        removed = [s for k, s in existing.items() if k not in kept_keys]
        if removed:
            kept = list(WorkflowStatus.objects.filter(project=project, key__in=kept_keys))
            fallback = next((s for s in kept if s.category == StatusCategory.NOT_STARTED), kept[0])
            done_fallback = next((s for s in kept if s.category == StatusCategory.DONE), fallback)
            for s in removed:
                repl = done_fallback if s.category == StatusCategory.DONE else fallback
                Task.objects.filter(project=project, status=s.key).update(status=repl.key)
                s.delete()

    return JsonResponse({"ok": True, "statuses": _status_json(project)})


# ---------------------------------------------------------------------------
# Saved baselines — multi-baseline capture & comparison (MS Project parity)
# ---------------------------------------------------------------------------

@require_POST
def baseline_save(request, project_id):
    """Capture current dates/progress as a new named baseline. Also refreshes each
    task's baseline_start/end so the Gantt baseline bar reflects the latest snapshot."""
    require_project_role(request, project_id, allowed_roles=EDIT_ROLES)
    project = get_object_or_404(Project, id=project_id, workspace=request.pmo_workspace)
    payload = _json_body(request) or {}
    name = (payload.get("name") or "").strip()[:80] or f"Baseline {project.baselines.count() + 1}"
    snapshot = {}
    with transaction.atomic():
        for t in project.tasks.all():
            snapshot[str(t.id)] = {
                "start": _iso(local_date(t.start_date)),
                "end": _iso(local_date(t.end_date)),
                "progress": t.progress,
            }
            t.baseline_start = t.start_date
            t.baseline_end = t.end_date
            t.save(update_fields=["baseline_start", "baseline_end"])
        b = Baseline.objects.create(project=project, name=name, snapshot=snapshot)
    return JsonResponse({"ok": True, "baseline": {
        "id": b.id, "name": b.name,
        "created_at": timezone.localtime(b.created_at).strftime("%d %b %Y %H:%M"),
        "snapshot": b.snapshot,
    }})


@require_POST
def baseline_delete(request, project_id, baseline_id):
    require_project_role(request, project_id, allowed_roles=EDIT_ROLES)
    project = get_object_or_404(Project, id=project_id, workspace=request.pmo_workspace)
    Baseline.objects.filter(id=baseline_id, project=project).delete()
    return JsonResponse({"ok": True})


# ---------------------------------------------------------------------------
# Per-resource capacity / PTO  +  automatic resource leveling
# ---------------------------------------------------------------------------

@require_POST
def resource_profile_save(request, project_id):
    """Upsert a workspace member's capacity (units) and time-off (PTO) ranges."""
    require_project_role(request, project_id, allowed_roles=EDIT_ROLES)
    project = get_object_or_404(Project, id=project_id, workspace=request.pmo_workspace)
    payload = _json_body(request) or {}
    try:
        user_id = int(payload.get("user_id"))
    except (TypeError, ValueError):
        return HttpResponseBadRequest("user_id is required.")
    if not WorkspaceMember.objects.filter(workspace=project.workspace, user_id=user_id).exists():
        return HttpResponseBadRequest("That person isn't a member of this workspace.")
    try:
        units = max(0.0, min(10.0, float(payload.get("units", 1.0))))
    except (TypeError, ValueError):
        units = 1.0
    try:
        rate = max(0.0, float(payload.get("rate", 0)))
    except (TypeError, ValueError):
        rate = 0.0
    wd = payload.get("working_days")
    working_days = None
    if isinstance(wd, list):
        working_days = sorted({int(x) for x in wd if isinstance(x, (int, float)) and 1 <= int(x) <= 7}) or None

    with transaction.atomic():
        profile, _ = ResourceProfile.objects.update_or_create(
            workspace=project.workspace, user_id=user_id,
            defaults={"units": units, "rate": rate, "working_days_json": working_days})
        profile.time_off.all().delete()
        for item in (payload.get("time_off") or []):
            s = _parse_dt(item.get("start"))
            e = _parse_dt(item.get("end"))
            if not s or not e:
                continue
            sd, ed = local_date(s), local_date(e)
            if ed < sd:
                sd, ed = ed, sd
            ResourceTimeOff.objects.create(
                profile=profile, start_date=sd, end_date=ed, note=(item.get("note") or "").strip()[:200] or None)
    return JsonResponse({"ok": True, "resources": _resources_json(project.workspace)})


def _resource_off_days(workspace):
    """Map user_id -> set(date) of PTO days across the workspace."""
    off = {}

    for profile in ResourceProfile.objects.filter(
        workspace=workspace
    ).prefetch_related("time_off"):

        days = set()

        for time_off in profile.time_off.all():
            current_date = time_off.start_date

            while current_date <= time_off.end_date:
                days.add(current_date)
                current_date += timedelta(days=1)

        if days:
            off[profile.user_id] = days

    return off


def _resource_working_days(workspace):
    """Map user_id -> set(ISO weekday) for people with a custom working week."""
    resource_working_days = {}

    for profile in ResourceProfile.objects.filter(workspace=workspace):
        if profile.working_days_json:
            resource_working_days[profile.user_id] = set(profile.working_days_json)

    return resource_working_days



def level_project_resources(project):
    """MS Project-style resource leveling: serialize each resource's overlapping
    tasks (respecting the project calendar, that person's PTO, and task
    dependencies) so nobody is booked on two things at once. Only ever delays
    tasks — never pulls them earlier — and returns a summary of the changes."""
    cal = get_calendar(project)
    off = _resource_off_days(project.workspace)
    rwd = _resource_working_days(project.workspace)
    tasks = list(project.tasks.all())
    child_ids = {t.parent_task_id for t in tasks if t.parent_task_id}
    # Leveling never moves manually-scheduled or inactive tasks.
    leaf = [t for t in tasks if t.id not in child_ids and t.start_date and t.end_date
            and t.is_active and t.scheduling_mode != SchedulingMode.MANUAL]
    if not leaf:
        return {"moved": 0, "finish_shift_days": 0, "new_finish": None}

    # Every resource assigned to each task (multi-assignment aware).
    task_users = {}
    for asg in TaskAssignment.objects.filter(task__project=project).values_list("task_id", "user_id"):
        task_users.setdefault(asg[0], []).append(asg[1])

    def users_for(t):
        return task_users.get(t.id) or ([t.assignee_id] if t.assignee_id else [])

    def working(uids, d):
        # a day works only if it's a project working day, and every assigned
        # resource is available that day (their own working week + no PTO)
        if not is_working_day(d, cal):
            return False
        for u in uids:
            if d in off.get(u, ()):
                return False
            wd = rwd.get(u)
            if wd and d.isoweekday() not in wd:
                return False
        return True

    def next_free(uids, d):
        for _ in range(1500):
            if working(uids, d):
                return d
            d += timedelta(days=1)
        return d

    def add_days(uids, start, n):
        n = max(1, n)
        d = next_free(uids, start)
        counted = 1
        while counted < n:
            d = next_free(uids, d + timedelta(days=1))
            counted += 1
        return d

    today = timezone.localtime(timezone.now()).date()
    p_start = local_date(project.start_date) or today
    orig_finish = max(local_date(t.end_date) for t in leaf)

    ordered = sorted(leaf, key=lambda t: (local_date(t.start_date) or p_start, t.sort_order, t.id))
    dates = {}
    resource_free = {}
    moved = 0

    for t in ordered:
        dur = count_working_days(local_date(t.start_date), local_date(t.end_date), cal)
        ns = local_date(t.start_date)
        dep_req = required_start_for(t, cal, dates)   # earliest allowed by predecessors/constraints
        if dep_req and dep_req > ns:
            ns = dep_req
        uids = users_for(t)
        if uids and not t.is_milestone:
            for u in uids:                            # no resource booked twice
                rf = resource_free.get(u)
                if rf and rf > ns:
                    ns = rf
            ns = next_free(uids, ns)
            ne = add_days(uids, ns, dur)
        else:
            ns = next_working_day(ns, cal)
            ne = add_working_days(ns, dur, cal)
        sd, ed = to_start_dt(ns), to_end_dt(ne)
        if sd != t.start_date or ed != t.end_date:
            moved += 1
            Task.objects.filter(id=t.id).update(start_date=sd, end_date=ed)
        dates[t.id] = (sd, ed)
        if uids and not t.is_milestone:
            for u in uids:
                resource_free[u] = next_free([u], ne + timedelta(days=1))

    new_finish = max(local_date(ed) for (_sd, ed) in dates.values())
    return {
        "moved": moved,
        "finish_shift_days": (new_finish - orig_finish).days,
        "new_finish": new_finish.isoformat(),
    }


@require_POST
def level_resources(request, project_id):
    require_project_role(request, project_id, allowed_roles=EDIT_ROLES)
    project = get_object_or_404(Project, id=project_id, workspace=request.pmo_workspace)
    with transaction.atomic():
        result = level_project_resources(project)
    return JsonResponse({"ok": True, **result})


# ---------------------------------------------------------------------------
# Portfolio roll-up dashboard — cross-project aggregation
# ---------------------------------------------------------------------------

def portfolio_rollup(request):
    """Each portfolio aggregated across its projects: weighted progress, tasks,
    overdue, milestones, open risks, span, resources and health."""
    workspace = request.pmo_workspace
    visible = list(_visible_projects(request).select_related("owner")
                   .prefetch_related("tasks", "raid_items", "memberships"))
    visible_ids = {p.id for p in visible}
    today = timezone.now()
    done_keys = set(
        WorkflowStatus.objects.filter(project__in=visible, category=StatusCategory.DONE)
        .values_list("key", flat=True)
    ) | {TaskStatus.DONE}

    def project_metrics(p):
        tasks = list(p.tasks.all())
        leaf = [t for t in tasks if not any(x.parent_task_id == t.id for x in tasks)]
        done = sum(1 for t in leaf if t.status in done_keys)
        overdue = sum(1 for t in leaf if t.end_date and t.end_date < today and t.status not in done_keys)
        starts = [t.start_date for t in leaf if t.start_date]
        ends = [t.end_date for t in leaf if t.end_date]
        return {
            "project": p,
            "tasks": len(leaf),
            "done": done,
            "overdue": overdue,
            "milestones": sum(1 for t in leaf if t.is_milestone),
            "progress": round(sum(t.progress for t in leaf) / len(leaf)) if leaf else 0,
            "open_raid": sum(1 for r in p.raid_items.all() if r.status != RaidStatus.CLOSED),
            "resources": len({t.assignee_id for t in leaf if t.assignee_id}),
            "start": min(starts) if starts else None,
            "end": max(ends) if ends else None,
            "health": "red" if overdue > 2 else ("amber" if overdue else "green"),
        }

    pm = {p.id: project_metrics(p) for p in visible}
    portfolios = (Portfolio.objects.filter(workspace=workspace).select_related("owner")
                  .prefetch_related("portfolio_projects__project"))
    cards, assigned_ids = [], set()
    for pf in portfolios:
        projs = [pp.project for pp in pf.portfolio_projects.all() if pp.project_id in visible_ids]
        assigned_ids.update(p.id for p in projs)
        if not projs and not _is_admin(request):
            continue
        rows = [pm[p.id] for p in projs]
        tot = sum(r["tasks"] for r in rows)
        weighted = round(sum(r["progress"] * r["tasks"] for r in rows) / tot) if tot else 0
        resources = set()
        for p in projs:
            resources.update(t.assignee_id for t in p.tasks.all() if t.assignee_id)
        starts = [r["start"] for r in rows if r["start"]]
        ends = [r["end"] for r in rows if r["end"]]
        overdue = sum(r["overdue"] for r in rows)
        cards.append({
            "portfolio": pf, "rows": rows, "project_count": len(projs),
            "tasks": tot, "done": sum(r["done"] for r in rows), "progress": weighted,
            "overdue": overdue, "milestones": sum(r["milestones"] for r in rows),
            "open_raid": sum(r["open_raid"] for r in rows), "resources": len(resources),
            "start": min(starts) if starts else None, "end": max(ends) if ends else None,
            "health": "red" if overdue > 3 else ("amber" if overdue else "green"),
        })

    unassigned = [pm[p.id] for p in visible if p.id not in assigned_ids]
    all_tasks = sum(r["tasks"] for r in pm.values())
    totals = {
        "portfolios": len(cards),
        "projects": len(visible),
        "tasks": all_tasks,
        "done": sum(r["done"] for r in pm.values()),
        "overdue": sum(r["overdue"] for r in pm.values()),
        "open_raid": sum(r["open_raid"] for r in pm.values()),
        "progress": round(sum(r["progress"] * r["tasks"] for r in pm.values()) / all_tasks) if all_tasks else 0,
    }
    return render(request, "pmo/portfolio_rollup.html", {
        "cards": cards, "unassigned": unassigned, "totals": totals,
    })


def portfolio_timeline(request):
    """Every visible project drawn as a single bar on one shared time axis —
    a portfolio-level Gantt, with milestone diamonds and today marker."""
    from datetime import date as _date

    visible = list(_visible_projects(request).select_related("owner")
                   .prefetch_related("tasks", "portfolios__portfolio"))
    today = timezone.now().date()
    done_keys = set(
        WorkflowStatus.objects.filter(project__in=visible, category=StatusCategory.DONE)
        .values_list("key", flat=True)
    ) | {TaskStatus.DONE}

    def as_date(v):
        if v is None:
            return None
        return v.date() if hasattr(v, "date") else v

    rows = []
    all_dates = []
    for p in visible:
        tasks = list(p.tasks.all())
        leaf = [t for t in tasks if not any(x.parent_task_id == t.id for x in tasks)]
        starts = [as_date(t.start_date) for t in leaf if t.start_date]
        ends = [as_date(t.end_date) for t in leaf if t.end_date]
        if not starts or not ends:
            continue
        p_start, p_end = min(starts), max(ends)
        overdue = sum(1 for t in leaf if t.end_date and as_date(t.end_date) < today and t.status not in done_keys)
        progress = round(sum(t.progress for t in leaf) / len(leaf)) if leaf else 0
        milestones = []
        for t in leaf:
            if t.is_milestone and t.start_date:
                milestones.append({"name": t.title, "date": as_date(t.start_date),
                                   "done": t.status in done_keys})
        pf = next((l.portfolio.name for l in p.portfolios.all()), None)
        rows.append({
            "project": p, "start": p_start, "end": p_end, "progress": progress,
            "overdue": overdue, "tasks": len(leaf), "milestones": milestones,
            "portfolio": pf,
            "health": "red" if overdue > 2 else ("amber" if overdue else "green"),
        })
        all_dates += [p_start, p_end]

    axis = {"months": [], "today_pct": None}
    if all_dates:
        lo, hi = min(all_dates), max(all_dates)
        # pad to month boundaries
        lo = lo.replace(day=1)
        hi_month_end = (hi.replace(day=1) + timedelta(days=32)).replace(day=1)
        span_days = max(1, (hi_month_end - lo).days)

        def pct(d):
            return round((d - lo).days / span_days * 100, 3)

        for r in rows:
            r["left_pct"] = pct(r["start"])
            r["width_pct"] = max(0.6, pct(r["end"]) - pct(r["start"]))
            for m in r["milestones"]:
                m["pct"] = pct(m["date"])
        # month gridlines
        cur = lo
        while cur < hi_month_end:
            axis["months"].append({"label": cur.strftime("%b %y"), "pct": pct(cur)})
            cur = (cur + timedelta(days=32)).replace(day=1)
        if lo <= today <= hi_month_end:
            axis["today_pct"] = pct(today)
        axis["start"], axis["end"] = lo, hi

    rows.sort(key=lambda r: (r["start"], r["end"]))
    totals = {
        "projects": len(rows),
        "milestones": sum(len(r["milestones"]) for r in rows),
        "overdue": sum(r["overdue"] for r in rows),
    }
    return render(request, "pmo/portfolio_timeline.html", {
        "rows": rows, "axis": axis, "totals": totals,
    })


def capabilities(request):
    """Overview of everything PMONexus can do (surfaced on the top bar)."""
    return render(request, "pmo/capabilities.html", {})


@require_POST
def task_attachment_upload(request, project_id, task_id):
    """Upload a real file (Word/PDF/email/image/…) as an attachment. Uses Azure
    Blob when configured, otherwise stores under MEDIA_ROOT for local dev."""
    import os as _os
    import uuid as _uuid
    from django.conf import settings

    require_project_role(request, project_id, allowed_roles=EDIT_ROLES)
    task = get_object_or_404(Task, id=task_id, project_id=project_id,
                             project__workspace=request.pmo_workspace)
    f = request.FILES.get("file")
    if not f:
        return HttpResponseBadRequest("No file uploaded.")
    safe_name = _os.path.basename(f.name)[:200] or "attachment"
    mime = f.content_type or "application/octet-stream"

    if _os.environ.get("AZURE_STORAGE_ACCOUNT_URL"):
        try:
            from storage.blob import upload_attachment
            url = upload_attachment(f"{task.project_id}/{task.id}/{_uuid.uuid4().hex}_{safe_name}", f, mime)
        except Exception as e:  # pragma: no cover
            return HttpResponseBadRequest(f"Upload failed: {e}")
    else:
        rel_dir = _os.path.join("attachments", str(task.project_id), str(task.id))
        dest_dir = _os.path.join(settings.MEDIA_ROOT, rel_dir)
        _os.makedirs(dest_dir, exist_ok=True)
        fname = f"{_uuid.uuid4().hex}_{safe_name}"
        with open(_os.path.join(dest_dir, fname), "wb") as out:
            for chunk in f.chunks():
                out.write(chunk)
        url = f"{settings.MEDIA_URL}{rel_dir}/{fname}".replace("\\", "/")

    a = TaskAttachment.objects.create(task=task, file_name=safe_name, blob_url=url, mime_type=mime)
    return JsonResponse({"ok": True, "id": a.id, "url": url, "name": safe_name, "mime": mime})


@require_POST
def task_create_recurring(request, project_id):
    """Create a recurring series of tasks (daily / weekly / monthly) grouped under
    a summary task — MS Project's Recurring Task."""
    require_project_role(request, project_id, allowed_roles=EDIT_ROLES)
    project = get_object_or_404(Project, id=project_id, workspace=request.pmo_workspace)
    payload = _json_body(request) or {}
    title = (payload.get("title") or "").strip()
    if not title:
        return HttpResponseBadRequest("Give the recurring task a name.")
    pattern = payload.get("pattern") or "WEEKLY"
    try:
        count = max(1, min(60, int(payload.get("count", 1))))
    except (TypeError, ValueError):
        count = 1
    try:
        duration = max(1, int(payload.get("duration", 1)))
    except (TypeError, ValueError):
        duration = 1
    cal = get_calendar(project)
    start0 = _parse_dt(payload.get("start_date"))
    base = (local_date(start0) if start0
            else (local_date(project.start_date) if project.start_date
                  else timezone.localtime(timezone.now()).date()))
    step = {"DAILY": timedelta(days=1), "WEEKLY": timedelta(weeks=1), "MONTHLY": timedelta(days=30)}.get(pattern, timedelta(weeks=1))
    sort_order = (project.tasks.aggregate(m=Max("sort_order"))["m"] or 0)
    created = []
    with transaction.atomic():
        sort_order += 1
        first = next_working_day(base, cal)
        parent = Task.objects.create(
            project=project, title=title, sort_order=sort_order,
            start_date=to_start_dt(first), end_date=to_end_dt(first))
        for i in range(count):
            s = next_working_day(base + step * i, cal)
            e = add_working_days(s, duration, cal)
            sort_order += 1
            created.append(Task.objects.create(
                project=project, parent_task=parent, title=f"{title} {i + 1}",
                start_date=to_start_dt(s), end_date=to_end_dt(e), sort_order=sort_order))
    return JsonResponse({"ok": True, "created": len(created)})
